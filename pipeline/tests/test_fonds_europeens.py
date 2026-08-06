"""Les fonds européens conventionnés : 7,88 Md€ programmés, zéro payé.

La fixture est la parution du 16 mars 2026 entière — 16 625 opérations —
réduite aux huit colonnes que le connecteur lit. Les quinze autres décrivent le
projet et son bénéficiaire ; les charger ferait un fichier de dix mégaoctets
sans rien ajouter à ce qui est vérifié ici.

Elle est stockée en CSV, lisible dans un diff, et rendue au classeur que le
connecteur lit vraiment : c'est le chemin de production — nom de feuille, date
d'arrêté, intitulés de colonnes, montants typés — qui est exercé, pas un lecteur
de test. Une seule chose se perd à la reconstruction : `openpyxl` écrit les
flottants à quinze chiffres significatifs, ce qui déplace le dernier bit de 15
montants sur 16 625. Les totaux sont donc vérifiés à l'euro, jamais au centime.
"""

import csv
import io
from pathlib import Path

import pytest

from plateforme import couverture
from plateforme.normalize import fonds_europeens as fe

FIXTURES = Path(__file__).parent / "fixtures"
FEUILLE = "LISTE OPERATION AU 16 03 2026"

# Un extrait de la page de référence, tel qu'elle sert le lien du fichier.
PAGE = """
  <section class="container" id="fichiers">
    <a href="https://www.europe-en-france.gouv.fr/sites/default/files/2026-03/\
20260316_liste_operations_conventionnees_FEDER_FSE_FTJ_0.xlsx" class="media fichier" download>
      <div class="title">20260316_liste_operations_conventionnees_FEDER_FSE_FTJ_0.xlsx</div>
    </a>
  </section>
"""


def _classeur() -> bytes:
    """La fixture CSV rendue au format XLSX de l'ANCT : deux feuilles, la
    notice et les données, et des cellules typées comme la source les type —
    les numéros d'opération purement numériques sont des nombres, les autres du
    texte, et c'est ce mélange qui interdit de traiter cette colonne comme une
    clé."""
    import openpyxl

    livre = openpyxl.Workbook(write_only=True)
    livre.create_sheet("INFO").append(["Date du document : 16/03/2026"])
    feuille = livre.create_sheet(FEUILLE)
    with (FIXTURES / "anct_feder_sample.csv").open(encoding="utf-8") as fichier:
        for rangee in csv.reader(fichier):
            feuille.append([_cellule(valeur) for valeur in rangee])
    tampon = io.BytesIO()
    livre.save(tampon)
    return tampon.getvalue()


def _cellule(valeur: str):
    if valeur == "":
        return None
    for type_ in (int, float):
        try:
            return type_(valeur)
        except ValueError:
            continue
    return valeur


@pytest.fixture(scope="module")
def lu():
    return fe.lire(_classeur())


@pytest.fixture(scope="module")
def operations(lu):
    return lu[0]


@pytest.fixture
def entrepot_anct(entrepot_seme):
    """L'entrepôt de production, plus le jeu de l'ANCT et deux régions.

    Le jeu et sa source viennent du registre semé — `infra/seed/` les porte.
    Deux régions suffisent à prouver que les codes du module joignent le
    référentiel — dont « 04 », que La Réunion écrit sur deux caractères et
    qu'une lecture en entier aurait transformé en « 4 ».
    """
    (connu,) = entrepot_seme.execute(
        "select count(*) from meta.dataset_registry where dataset_id = ?",
        (fe.DATASET,),
    ).fetchone()
    assert connu == 1, "le jeu doit être déclaré dans infra/seed/dataset_registry.csv"
    for code, nom in (("04", "La Réunion"), ("84", "Auvergne-Rhône-Alpes")):
        entrepot_seme.execute(
            "insert into geo.geography_reference (geo_level, geo_code, vintage,"
            " name, parent_level, parent_code, flags) values ('region', ?, ?, ?,"
            " null, null, '{}')",
            (code, fe.MILLESIME, nom),
        )
    entrepot_seme.commit()
    return entrepot_seme


def test_le_lien_du_fichier_se_retrouve_dans_la_page():
    """L'URL porte la date de parution et change à chaque mise à jour ; la page
    de référence, elle, est stable."""
    assert fe.trouver_url(PAGE).endswith(
        "2026-03/20260316_liste_operations_conventionnees_FEDER_FSE_FTJ_0.xlsx"
    )


def test_entre_deux_parutions_c_est_la_plus_recente_qui_est_prise():
    """La page garde parfois la parution précédente à côté de la nouvelle : les
    départager par la date du nom de fichier, et non par l'ordre d'apparition."""
    ancienne = PAGE.replace("2026-03/20260316", "2025-09/20250912")
    assert "20250912" in ancienne
    assert "20260316" in fe.trouver_url(ancienne + PAGE)
    assert "20260316" in fe.trouver_url(PAGE + ancienne)


def test_une_page_sans_fichier_ne_charge_rien():
    """Recharger éternellement une liste périmée sans que rien ne le dise est le
    seul défaut qu'une URL codée en dur produirait."""
    with pytest.raises(ValueError, match="a renommé son fichier"):
        fe.trouver_url("<html><a href='autre-chose.pdf'>ailleurs</a></html>")


def test_la_parution_lue_est_celle_qui_est_annoncee(lu):
    operations, arrete = lu
    assert arrete == "2026-03-16"
    assert len(operations) == fe.OPERATIONS == 16_625


def test_les_chiffres_de_la_parution(operations):
    """Les deux contrôles bloquants passent, et rendent les repères du jeu."""
    verifies = fe.controler(operations)
    assert verifies["operations"] == 16_625
    assert verifies["lignes_verifiees"] == 16_591
    assert verifies["montant_ue_absent"] == 34
    assert round(verifies["montant_ue_total"]) == 7_879_820_895
    assert round(verifies["depenses_eligibles_total"]) == 17_952_782_937
    assert verifies["programmes"] == 21
    assert verifies["volet_national"] == 398
    # Trois fonds, une masse : le FSE+ n'est pas loin du FEDER, et le FTJ pèse
    # dix fois moins que l'un comme l'autre.
    fonds = {nom: round(montant / 1e6) for nom, montant in verifies["montant_ue_par_fonds"].items()}
    assert fonds == {"FEDER": 4_048, "FSE+": 3_379, "FTJ": 453}
    # Aucun montant négatif : une aide européenne ne se reprend pas ici.
    assert all(operation["ue"] is None or operation["ue"] >= 0 for operation in operations)


def test_le_montant_ue_se_recalcule_sur_les_lignes_renseignees(operations):
    """Dépenses éligibles × taux = montant UE. C'est le seul contrôle qui puisse
    voir une colonne qui aurait glissé : il compare trois colonnes entre elles,
    là où un total ne se compare qu'à lui-même."""
    verifies = fe.controler(operations)
    assert verifies["ecart_relatif_maximal"] < 0.0001
    assert verifies["ecart_relatif_maximal"] < fe.TOLERANCE


def test_le_recalcul_bloque_sur_un_montant_decale(operations):
    faussee = dict(operations[0], ue=operations[0]["ue"] * 1.02)
    with pytest.raises(fe.ConventionnementRompu, match="d'écart"):
        fe.controler([faussee, *operations[1:]])


def test_le_decompte_bloque_une_lecture_tronquee(operations):
    """Un classeur lu à moitié rend une somme parfaitement cohérente et
    parfaitement fausse : aucun total interne ne peut le voir."""
    with pytest.raises(fe.ConventionnementRompu, match="16624 opérations lues"):
        fe.controler(operations[:-1])
    with pytest.raises(fe.ConventionnementRompu, match="contre 16626 attendues"):
        fe.controler(operations, attendu=16_626)


def test_un_programme_inconnu_bloque(operations):
    """Sans code région, ses opérations disparaîtraient de la carte sans manquer
    à aucun total : c'est exactement le défaut qu'aucun contrôle de masse ne
    voit."""
    intrus = dict(operations[0], programme="Programme Atlantide FEDER 2028-2034")
    with pytest.raises(fe.ConventionnementRompu, match="Atlantide"):
        fe.controler([intrus, *operations[1:]])


def test_la_correspondance_couvre_exactement_les_programmes_du_fichier(operations):
    assert {operation["programme"] for operation in operations} == set(fe.PROGRAMMES)
    codes = {code for code in fe.PROGRAMMES.values() if code}
    assert len(codes) == 18, "dix-huit régions, une par programme régional"
    assert all(len(code) == 2 for code in codes), "codes INSEE sur deux caractères"


def test_la_colonne_region_de_la_source_est_bien_inutilisable(operations):
    """La raison pour laquelle la région vient du programme : 36,9 % de lignes
    vides et 71 orthographes pour 18 régions."""
    vides = [o for o in operations if not o["region_source"]]
    assert len(vides) == 6_135
    assert 0.368 < len(vides) / len(operations) < 0.370
    modalites = {o["region_source"] for o in operations}
    assert len(modalites) == 71
    assert {"La Réunion", "04/La Réunion"} <= modalites, "deux écritures, un territoire"
    # Le programme, lui, est servi partout.
    assert all(operation["programme"] for operation in operations)


def test_la_region_vient_du_programme(operations):
    candidates = fe.par_region(operations)
    assert len(candidates) == 18
    montants = {code: montant for _, _, code, _, montant in candidates}
    # La Réunion conventionne plus que l'Auvergne-Rhône-Alpes : 715 M€ contre
    # 668 M€ pour vingt fois moins d'habitants. C'est ce que ce jeu montre.
    assert round(montants["04"]) == 715_328_854
    assert round(montants["84"]) == 667_934_631
    assert round(montants["32"]) == 653_158_812
    assert round(montants["44"]) == 574_375_464
    assert all(periode == "2021-2027" for _, _, _, periode, _ in candidates)


def test_ce_qui_n_est_pas_regionalisable_n_est_pas_reparti(operations):
    """Les deux programmes nationaux et Saint-Martin sortent de la maille
    régionale : 2,34 Md€ nommés plutôt que ventilés au prorata de rien."""
    hors = fe.non_regionalise(operations)
    national = hors["Programme national FSE+ Emploi - Inclusion - Jeunesse - Compétences"]
    assert national["operations"] == 7_803
    assert round(national["montant_ue"]) == 2_127_915_705
    assert hors["Programme national FTJ Emploi - Compétences"]["operations"] == 209
    assert hors["Programme Saint Martin FEDER 2021-2027"]["operations"] == 20
    assert len(hors) == 3
    # Les 398 opérations « Volet national » sont toutes portées par le programme
    # national, donc déjà hors maille : aucun motif « volet national » n'apparaît
    # dans cette parution. Le filtre reste pour le jour où un programme régional
    # en portera une — auquel cas elle sera écartée, pas attribuée à sa région.
    assert not [motif for motif in hors if motif.startswith("volet national")]
    volet = [o for o in operations if o["region_source"] == fe.VOLET_NATIONAL]
    assert len(volet) == 398
    assert {fe.PROGRAMMES[o["programme"]] for o in volet} == {None}

    regionalise = sum(montant for _, _, _, _, montant in fe.par_region(operations))
    total = fe.controler(operations)["montant_ue_total"]
    assert round(regionalise) == 5_542_014_566
    assert round(total - regionalise) == 2_337_806_329


def test_le_numero_d_operation_n_est_pas_une_cle(operations):
    """Le n° 54030 désigne deux opérations différentes, l'une en Bretagne,
    l'autre en Nouvelle-Aquitaine : la clé est (NUMCCI, numéro), et même là deux
    lignes du programme national sont strictement identiques."""
    numeros = [o["operation"] for o in operations]
    assert len(numeros) - len(set(numeros)) == 5
    collision = [o for o in operations if o["operation"] == "54030"]
    assert {o["programme"] for o in collision} == {
        "Programme Bretagne FEDER-FSE+ 2021-2027",
        "Programme FEDER FSE+ Nouvelle-Aquitaine",
    }
    assert fe.controler(operations)["cles_dupliquees"] == 2


def test_le_bruit_flottant_est_arrondi_au_centime(operations):
    """Le tableur rend 49 999,99999999999 là où la convention dit 50 000."""
    bretagne = next(
        o for o in operations
        if o["operation"] == "54030" and o["cci"] == "2021FR16FFPR016"
    )
    assert bretagne["ue"] == 50_000.0
    assert all(
        o["ue"] is None or round(o["ue"], 2) == o["ue"] for o in operations
    )


def test_les_fiches_disent_conventionne_et_non_paye():
    assert len(fe.FICHE["public"].split()) <= 50
    assert "pas un montant payé" in fe.FICHE["public"]
    assert "programmés, et non des montants payés" in fe.FICHE["technique"]
    assert "Volet national" in fe.FICHE["technique"]
    assert "ne sont donc répartis sur aucune région" in fe.FICHE["technique"]
    assert "Libellé Programme" in fe.FICHE["technique"]


def test_declarer_puis_redeclarer_contre_un_vrai_entrepot(entrepot_anct):
    fe.declarer(entrepot_anct)
    fe.declarer(entrepot_anct)
    publies = entrepot_anct.execute(
        "select indicator_id, unit, additive, theme, time_granularity"
        " from core.indicators where dataset_id = ?",
        (fe.DATASET,),
    ).fetchall()
    assert publies == [
        (fe.INDICATEUR, "EUR", True, "europe", "pluriannuelle")
    ]
    (orphelines,) = entrepot_anct.execute(
        "select count(*) from core.indicator_definitions d where not exists"
        " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
    ).fetchone()
    assert orphelines == 0


def test_l_ecriture_reelle_contre_un_vrai_entrepot(entrepot_anct, operations):
    from plateforme import entrepot, revisions
    from plateforme.normalize.ofgl import filtrer_territoires_connus

    fe.declarer(entrepot_anct)
    run_id = entrepot.start_run(entrepot_anct, fe.DATASET, "manual")
    candidates = fe.par_region(operations)
    gardees, ecartes = filtrer_territoires_connus(entrepot_anct, candidates)
    ecrites, revisees = revisions.remplacer(
        entrepot_anct, run_id, [fe.INDICATEUR], fe.COLONNES,
        [
            (identifiant, niveau, code, fe.MILLESIME, periode, montant)
            for identifiant, niveau, code, periode, montant in gardees
        ],
    )
    entrepot_anct.commit()
    assert (ecrites, revisees) == (2, 0), "deux régions au référentiel, rien à réviser"
    assert len(ecartes) == 16, "les seize autres régions n'y sont pas encore"
    (valeur,) = entrepot_anct.execute(
        "select value from core.observations where indicator_id = ?"
        " and geo_code = '04' and period = '2021-2027'",
        (fe.INDICATEUR,),
    ).fetchone()
    assert round(valeur) == 715_328_854

    # Une révision se voit : la valeur d'hier est archivée, pas écrasée.
    _, revisees = revisions.remplacer(
        entrepot_anct, run_id, [fe.INDICATEUR], fe.COLONNES,
        [(fe.INDICATEUR, "region", "04", fe.MILLESIME, fe.PERIODE, 1.0),
         (fe.INDICATEUR, "region", "84", fe.MILLESIME, fe.PERIODE, 667_934_630.9)],
    )
    entrepot_anct.commit()
    assert revisees == 1, "seule La Réunion a bougé"


def test_la_couverture_se_mesure_en_euros(operations):
    candidates = fe.par_region(operations)
    lus = fe.montant_par_maille(candidates)
    assert round(lus["region"]) == 5_542_014_566
    couverture.controler(lus, lus)
    with pytest.raises(couverture.CouvertureInsuffisante):
        couverture.controler(lus, {"region": lus["region"] * 0.9})
