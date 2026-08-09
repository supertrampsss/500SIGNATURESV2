"""Les subventions de l'État aux associations : deux codes commune pour un SIRET.

Le jaune budgétaire porte lui-même un code commune par versement, produit par
l'appariement Sirene de l'administration en 2022. Le répertoire Sirene
d'aujourd'hui en donne un autre pour le même SIRET. Les confronter est le seul
contrôle qui puisse voir une colonne décalée ou un SIRET reconstruit sans ses
zéros de tête : la somme du fichier, elle, reste égale à elle-même dans les deux
cas.

La fixture est donc la parution entière — 102 619 versements — réduite aux cinq
colonnes que le connecteur lit, augmentée des deux colonnes que la requête
Sirene a rendues. **La correspondance y est déjà résolue** : aucun test ne
touche au réseau, `resoudre()` n'est jamais appelée ici.

Rien ne se perd à la reconstruction. Les montants du jaune font tous exactement
dix-huit caractères, cadrés à droite par des espaces insécables ; la fixture les
stocke détourés et le classeur de test les recadre, si bien que le connecteur
lit la chaîne exacte de la source — rembourrage insécable, séparateur de
milliers insécable et virgule décimale compris.
"""

import csv
import io
import json
import math
from collections import defaultdict
from pathlib import Path

import pytest

from plateforme import couverture
from plateforme.normalize import subventions as s

FIXTURES = Path(__file__).parent / "fixtures"

# Les montants du jaune sont cadrés à droite sur cette largeur, avec des espaces
# insécables. Les 102 619 lignes la respectent sans exception : c'est ce qui
# permet à la fixture de les stocker détourés sans rien perdre.
LARGEUR_MONTANT = 18

COLONNES = [
    "Programme", "SIREN", "NIC", "Montant", "Dénomination", "Objet 2021", "COG : code",
]

# Les colonnes que le tableur rend en texte. Sans elles, une dénomination faite
# de chiffres — il en existe — deviendrait un entier et perdrait ses zéros de
# tête, exactement le défaut que la reconstruction cherche à exercer ailleurs.
TEXTUELLES = {"Dénomination", "Objet 2021"}


def _masse(lignes) -> float:
    return math.fsum(ligne["montant"] for ligne in lignes)


def _cellule(colonne: str, valeur: str):
    """La cellule telle que le tableur la rend, à partir du texte de la fixture.

    Les types comptent : le tableur rend `Programme`, `SIREN`, `NIC` et
    `COG : code` en nombres — ce qui mange les zéros de tête —, sauf le SIREN
    absent (« N/A ») et les codes corses (2A004), qui restent du texte. Rendre
    tout en chaînes ferait passer un connecteur qui ne rembourre rien.
    """
    if colonne == "Montant":
        return valeur.rjust(LARGEUR_MONTANT, "\xa0")
    if colonne == "NIC":
        return int(valeur) if valeur else None
    if colonne in TEXTUELLES:
        return valeur or None
    return int(valeur) if valeur.isdigit() else valeur


def _source() -> tuple[bytes, dict[str, tuple[str, str]]]:
    """-> (classeur du jaune, correspondance Sirene déjà résolue).

    Le classeur est reconstruit tel que la source le publie : une feuille
    nommée, un titre et une ligne vide avant l'en-tête. C'est le chemin de
    production qui est exercé — nom de feuille, intitulés de colonnes, types de
    cellules — et non un lecteur de test.
    """
    import openpyxl

    livre = openpyxl.Workbook(write_only=True)
    feuille = livre.create_sheet(s.FEUILLE)
    feuille.append(["Crédits attribués au monde associatif en 2021"])
    feuille.append([])
    feuille.append(COLONNES)
    sirene = {}
    with (FIXTURES / "plf2023_subventions_associations_sample.csv").open(
        encoding="utf-8"
    ) as fichier:
        for rang in csv.DictReader(fichier):
            cellules = {colonne: _cellule(colonne, rang[colonne]) for colonne in COLONNES}
            feuille.append([cellules[colonne] for colonne in COLONNES])
            # Un SIRET connu de Sirene mais sans commune y entre quand même, la
            # colonne d'état faisant foi : c'est un cas que le contrôle doit
            # compter comme non résolu, pas comme absent du répertoire.
            if rang["sirene_etat"]:
                siret = s._siret(cellules["SIREN"], cellules["NIC"])
                sirene[siret] = (rang["sirene_commune"], rang["sirene_etat"])
    tampon = io.BytesIO()
    livre.save(tampon)
    return tampon.getvalue(), sirene


@pytest.fixture(scope="module")
def source():
    return _source()


@pytest.fixture(scope="module")
def lignes(source):
    return s.lire(source[0])


@pytest.fixture(scope="module")
def sirene(source):
    return source[1]


def test_les_espaces_insecables_ne_se_lisent_pas_toutes_seules():
    """Le piège du fichier. `float()` lève sur la chaîne du tableur, et retirer
    l'espace ordinaire ne change rien : c'est l'insécable qui sépare les
    milliers."""
    brut = "\xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa0\xa03\xa0850,00"
    with pytest.raises(ValueError):
        float(brut)
    with pytest.raises(ValueError):
        float(brut.replace(" ", "").replace(",", "."))
    assert s._nombre(brut) == 3850.0
    assert s._nombre("527\xa0052\xa0389,14") == 527_052_389.14
    assert s._nombre("-10\xa0987,50") == -10987.5
    assert s._nombre(None) is None


def test_les_zeros_de_tete_manges_par_le_tableur_sont_rendus():
    """La colonne commune et les deux colonnes du SIRET sont typées en nombres
    par la source. Sans rembourrage, 5 248 versements tomberaient hors
    référentiel et 55 500 SIRET ne trouveraient rien dans Sirene — sans qu'aucun
    total ne bouge."""
    assert s._commune(1004) == "01004"
    assert s._commune("2A004") == "2A004"
    assert s._siret(301294831, 88) == "30129483100088"
    assert s._siret(57817314, 32) == "05781731400032"
    assert s._siret("N/A", None) == ""


def test_le_classeur_entier_se_lit(lignes):
    assert len(lignes) == 102_619
    # Le document entier est composé d'espaces insécables — « APSIS\xa0EMERGENCE ».
    # Publié tel quel, un nom se recopie mal, ne se cherche pas et coupe au
    # mauvais endroit : la lecture les ramène à l'espace ordinaire.
    assert lignes[0] == {
        "programme": "101",
        "siren": "301294831",
        "siret": "30129483100088",
        "montant": 3850.0,
        "commune": "57672",
        "nom": "APSIS EMERGENCE",
        "objet": "Transfert direct aux ménages qui transitent par une association",
    }
    assert round(math.fsum(ligne["montant"] for ligne in lignes), 2) == 10_543_390_140.84
    # Quatre reversements : des lignes négatives que rien n'annonce, et qu'un
    # `abs()` posé par confort ferait disparaître du total.
    negatifs = [ligne for ligne in lignes if ligne["montant"] < 0]
    assert len(negatifs) == 4
    assert round(math.fsum(ligne["montant"] for ligne in negatifs), 2) == -22_674.41


def test_le_classeur_refuse_de_changer_de_forme(source):
    """La feuille et les sept colonnes sont contrôlées, pas supposées."""
    import openpyxl

    livre = openpyxl.load_workbook(io.BytesIO(source[0]))
    livre[s.FEUILLE].title = "versements 2022"
    tampon = io.BytesIO()
    livre.save(tampon)
    with pytest.raises(ValueError, match="absente du classeur"):
        s.lire(tampon.getvalue())


def test_le_code_commune_du_jaune_se_retrouve_dans_sirene(lignes, sirene):
    """Le contrôle bloquant, sur les chiffres mesurés.

    Deux appariements indépendants, à cinq ans d'écart : celui de
    l'administration en 2022 et le répertoire du 1er août 2026. Ils tombent sur
    la même commune pour 99,974 % des euros résolus."""
    verifies = s.controler(lignes, sirene)
    assert verifies["versements"] == 102_619
    assert verifies["programmes"] == 110
    assert verifies["etablissements"] == 55_500
    assert verifies["sans_siret"] == 660
    assert round(verifies["taux_resolution_euros"], 6) == 0.999670
    assert round(verifies["taux_concordance_euros"], 6) == 0.999743
    assert verifies["desaccords"] == 64
    assert verifies["masse_desaccords"] == 2_559_841.61
    assert s.RESOLUTION_MINIMALE < verifies["taux_resolution_euros"]
    assert s.CONCORDANCE_MINIMALE < verifies["taux_concordance_euros"]


def test_sirene_dit_ce_que_le_jaune_tait(lignes, sirene):
    """La colonne « Etat administratif » du jaune vaut « Non déterminé » sur ses
    102 619 lignes. Sirene, lui, sait que plus d'un versement sur cinq est allé
    à un établissement qui n'existe plus."""
    verifies = s.controler(lignes, sirene)
    assert verifies["versements_etablissement_ferme"] == 21_181
    assert verifies["masse_etablissement_ferme"] == 1_548_640_174.98
    assert 0.20 < verifies["versements_etablissement_ferme"] / len(lignes) < 0.21


def test_le_controle_bloque_si_les_siret_perdent_leurs_zeros(lignes, sirene):
    """Le défaut que ce contrôle existe pour attraper : `siren + nic` sans
    rembourrage. Le total du fichier ne bouge pas d'un centime, et plus rien ne
    se retrouve dans le répertoire."""
    fausses = [
        {**ligne, "siret": ligne["siret"][:9] + ligne["siret"][9:].lstrip("0")}
        for ligne in lignes
    ]
    assert _masse(fausses) == _masse(lignes)
    with pytest.raises(s.ConcordanceRompue, match="SIRET du jaune"):
        s.controler(fausses, sirene)


def test_le_controle_bloque_si_la_colonne_commune_glisse(lignes, sirene):
    """Une colonne décalée d'une ligne : même nombre de versements, même total,
    des codes commune tous valides — et les établissements placés ailleurs."""
    decalees = [
        {**ligne, "commune": suivante["commune"]}
        for ligne, suivante in zip(lignes, lignes[1:], strict=False)
    ]
    decalees.append(lignes[-1])
    assert len(decalees) == len(lignes)
    assert _masse(decalees) == _masse(lignes)
    with pytest.raises(s.ConcordanceRompue, match="ailleurs que Sirene"):
        s.controler(decalees, sirene)


def test_le_controle_bloque_sur_une_lecture_tronquee(lignes, sirene):
    """Un classeur lu à moitié rend un total parfaitement cohérent : seul le
    décompte de la parution peut le voir."""
    with pytest.raises(s.ConcordanceRompue, match="attendus"):
        s.controler(lignes[:-1], sirene)
    with pytest.raises(s.ConcordanceRompue, match="est vide"):
        s.controler([], sirene)
    with pytest.raises(s.ConcordanceRompue, match="clé vers Sirene"):
        s.controler([{**ligne, "siret": ""} for ligne in lignes], sirene)


def test_les_arrondissements_se_replient_sur_leur_commune(lignes):
    """Le jaune ne code Paris, Lyon et Marseille que par arrondissement. Les
    laisser tels quels publierait Paris à zéro — le défaut que `couverture.py`
    raconte pour le répertoire des logements sociaux."""
    candidates, _ = s.par_commune(lignes)
    montants = {
        code: valeur
        for identifiant, _, code, _, valeur in candidates
        if identifiant == "etat_subventions_associations"
    }
    assert not {code for code in montants if s.commune_mere(code)}
    assert montants["75056"] == 3_232_806_906.55
    assert montants["69123"] == 178_630_932.97
    assert montants["13055"] == 168_819_395.20


def test_paris_pese_un_tiers_parce_que_les_sieges_y_sont(lignes):
    """La réserve que ce jeu existe pour porter : la commune est celle de
    l'établissement bénéficiaire, pas celle où l'action est menée."""
    candidates, _ = s.par_commune(lignes)
    montants = {
        code: valeur
        for identifiant, _, code, _, valeur in candidates
        if identifiant == "etat_subventions_associations"
    }
    total = math.fsum(montants.values())
    assert 0.32 < montants["75056"] / total < 0.33
    dix = sorted(montants.values(), reverse=True)[:10]
    assert 0.43 < math.fsum(dix) / total < 0.44


def test_le_denombrement_donne_son_denominateur_au_montant(lignes):
    """Trois milliards pour une association ou pour 3 842 ne se lisent pas
    pareil : les deux indicateurs sortent du même passage."""
    candidates, _ = s.par_commune(lignes)
    etablissements = {
        code: valeur
        for identifiant, _, code, _, valeur in candidates
        if identifiant == "etat_subventions_associations_etablissements"
    }
    assert etablissements["75056"] == 3842
    assert etablissements["33063"] == 568
    assert sum(etablissements.values()) == 55_351
    # Un code commune porte les deux indicateurs ou aucun : sinon un montant
    # s'afficherait sans son dénominateur.
    montants = {
        code for identifiant, _, code, _, _ in candidates
        if identifiant == "etat_subventions_associations"
    }
    assert set(etablissements) == montants


def test_l_etranger_et_l_outre_mer_sont_ecartes_pas_repartis(lignes):
    """La plus grosse ligne du fichier — 527,1 M€ à l'Association internationale
    de développement, aux États-Unis — n'est pas une subvention communale. La
    diluer fabriquerait de l'argent que personne n'a reçu."""
    candidates, ecartes = s.par_commune(lignes)
    assert ecartes == {
        "collectivite_outre_mer": {"versements": 727, "montant": 31_654_387.91},
        "pays_etranger": {"versements": 164, "montant": 569_118_677.05},
    }
    codes = {code for _, _, code, _, _ in candidates}
    assert not {code for code in codes if s.hors_territoire(code)}
    # Rien ne se perd en chemin : ce qui est publié plus ce qui est écarté
    # nommément redonne le fichier, au centime.
    ecarte = math.fsum(motif["montant"] for motif in ecartes.values())
    assert round(s.masse_par_maille(candidates)["commune"] + ecarte, 2) == (
        round(math.fsum(ligne["montant"] for ligne in lignes), 2)
    )


def test_la_couverture_se_mesure_en_euros(lignes):
    """8 716 codes communaux candidats pour 9,94 Md€. La couverture compte des
    euros et non des codes : treize communes disparues du COG depuis 2021 ne
    pèsent pas treize huit-millièmes du total, elles pèsent 1,76 M€."""
    candidates, _ = s.par_commune(lignes)
    lus = s.masse_par_maille(candidates)
    assert lus == {"commune": 9_942_617_075.88}
    # Le dénombrement n'entre pas dans la mesure : mélanger des euros et des
    # établissements ferait une part que rien ne mesure.
    assert len(lus) == 1
    couverture.controler(lus, lus)
    with pytest.raises(couverture.CouvertureInsuffisante):
        couverture.controler(lus, {"commune": lus["commune"] * 0.5})


def test_l_url_de_sirene_est_retrouvee_dans_le_catalogue():
    """L'URL du parquet porte la date de production et change chaque mois. La
    coder en dur ferait vérifier le jaune contre un répertoire périmé sans que
    la concordance ne bouge — le contrôle passerait, il ne contrôlerait plus."""
    catalogue = json.dumps({"resources": [
        {"title": "Sirene : Fichier StockUniteLegale - 01 août 2026 (format parquet)",
         "format": "parquet", "url": "https://static.data.gouv.fr/x/unitelegale.parquet"},
        {"title": "Sirene : Fichier StockEtablissement - 01 août 2026 (format parquet)",
         "format": "parquet", "url": "https://static.data.gouv.fr/x/etablissement.parquet"},
        {"title": "Sirene : Fichier StockEtablissementHistorique - 01 août 2026"
                  " (format parquet)",
         "format": "parquet", "url": "https://static.data.gouv.fr/x/historique.parquet"},
        {"title": "Sirene : Fichier StockEtablissementLiensSuccession - 01 août 2026"
                  " (format parquet)",
         "format": "parquet", "url": "https://static.data.gouv.fr/x/succession.parquet"},
        {"title": "Sirene : Fichier StockEtablissement - 01 août 2026 (format csv)",
         "format": "csv", "url": "https://static.data.gouv.fr/x/etablissement.zip"},
    ]})
    assert s.url_sirene(catalogue) == "https://static.data.gouv.fr/x/etablissement.parquet"
    with pytest.raises(ValueError, match="une seule ressource"):
        s.url_sirene(json.dumps({"resources": []}))


def test_les_fiches_portent_les_reserves():
    """Quatre réserves, dont aucune ne se devine à la lecture d'un montant."""
    montant = s.INDICATEURS["etat_subventions_associations"][1]
    # L'exercice a cinq ans, et la fiche publique le dit elle-même.
    assert "2021" in montant
    # La commune est celle du bénéficiaire, pas celle de l'action.
    assert "l'établissement bénéficiaire" in montant
    assert "pas celle où l'action est menée" in montant
    # L'État, pas les collectivités.
    assert "versements de l'État, pas des collectivités" in montant
    # Des établissements ont fermé depuis : la fiche technique le chiffre.
    assert "aujourd'hui fermés" in s.TECHNIQUE
    assert "l'établissement n'existe plus" in s.TECHNIQUE
    assert "n'y sont pas" in s.TECHNIQUE and "collectivités" in s.TECHNIQUE
    assert "ne sont répartis sur aucune commune" in s.TECHNIQUE
    for identifiant, (_, publique, _, _) in s.INDICATEURS.items():
        assert len(publique.split()) <= 50, identifiant


def _semer_le_jeu(conn) -> None:
    """Le jeu au registre : sans lui, l'indicateur pend à un `dataset_id` que
    `entrepot.verifier_integrite` compte comme orphelin, et la clé étrangère du
    run refuse de démarrer."""
    conn.execute(
        "insert or ignore into meta.dataset_registry (dataset_id, source_id, title,"
        " priority) values (?, ?, 'PLF 2023 - effort financier de l''Etat en faveur"
        " des associations', 'P2')",
        (s.DATASET, s.SOURCE),
    )


def test_declarer_puis_redeclarer_contre_un_vrai_entrepot(entrepot_seme):
    _semer_le_jeu(entrepot_seme)
    s.declarer(entrepot_seme)
    s.declarer(entrepot_seme)
    publies = dict(entrepot_seme.execute(
        "select indicator_id, unit from core.indicators where dataset_id = ?",
        (s.DATASET,),
    ).fetchall())
    assert publies == {
        "etat_subventions_associations": "EUR",
        "etat_subventions_associations_etablissements": "count",
    }
    # Un dénombrement n'est ni en euros courants ni dans une comptabilité
    # budgétaire : recopier les colonnes du montant mentirait sur sa nature.
    cadres = dict(entrepot_seme.execute(
        "select indicator_id, coalesce(price_basis, '') || '/' ||"
        " coalesce(accounting_frame, '') from core.indicators where dataset_id = ?",
        (s.DATASET,),
    ).fetchall())
    assert cadres["etat_subventions_associations"] == "current/budgetaire"
    assert cadres["etat_subventions_associations_etablissements"] == "/"
    (orphelines,) = entrepot_seme.execute(
        "select count(*) from core.indicator_definitions d where not exists"
        " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
    ).fetchone()
    assert orphelines == 0


def test_l_ecriture_reelle_contre_un_vrai_entrepot(entrepot_seme, lignes):
    from plateforme import entrepot, revisions
    from plateforme.normalize.ofgl import filtrer_territoires_connus

    _semer_le_jeu(entrepot_seme)
    s.declarer(entrepot_seme)
    for code, nom in (("75056", "Paris"), ("33063", "Bordeaux")):
        entrepot_seme.execute(
            "insert into geo.geography_reference (geo_level, geo_code, vintage, name,"
            " parent_level, parent_code, flags) values ('commune', ?, ?, ?,"
            " 'departement', ?, '{}')",
            (code, s.MILLESIME, nom, code[:2]),
        )
    run_id = entrepot.start_run(entrepot_seme, s.DATASET, "manual")
    candidates, _ = s.par_commune(lignes)
    gardees, ecartes = filtrer_territoires_connus(entrepot_seme, candidates)
    ecrites, _ = revisions.remplacer(
        entrepot_seme, run_id, sorted(s.INDICATEURS), s.COLONNES,
        [
            (identifiant, niveau, code, s.MILLESIME, periode, valeur)
            for identifiant, niveau, code, periode, valeur in gardees
        ],
    )
    entrepot_seme.commit()
    assert ecrites == 4, "deux indicateurs, deux communes, un exercice"
    assert len(ecartes) > 0, "les 8 714 autres communes ne sont pas au référentiel"
    (montant,) = entrepot_seme.execute(
        "select value from core.observations"
        " where indicator_id = 'etat_subventions_associations'"
        " and geo_code = '75056' and period = ?",
        (s.EXERCICE,),
    ).fetchone()
    assert montant == 3_232_806_906.55
    (etablissements,) = entrepot_seme.execute(
        "select value from core.observations"
        " where indicator_id = 'etat_subventions_associations_etablissements'"
        " and geo_code = '33063' and period = ?",
        (s.EXERCICE,),
    ).fetchone()
    assert etablissements == 568


# --------------------------------------------------------------- les noms
#
# « 353 € par habitant » ne répond pas à la question qu'on vient poser — *à
# quelles associations* — et un total sans ses bénéficiaires demande de croire
# sur parole. Ce document existe pour être lu bénéficiaire par bénéficiaire.


def test_les_beneficiaires_sont_nommes_et_sommes_par_programme(lignes):
    nommes = s.beneficiaires(lignes)
    # Le grain est le couple (bénéficiaire, programme) : une même association
    # reçoit plusieurs versements au titre d'un même programme, et les sommer
    # ici évite qu'un total dépende de qui l'additionne.
    assert len(nommes) < len(lignes)
    cles = [(code, siren, programme) for code, siren, _, programme, _, _ in nommes]
    assert len(cles) == len(set(cles)), "un couple (bénéficiaire, programme) apparaît deux fois"
    # APSIS EMERGENCE, Thionville, programme 101 : 3 850 + 33 330 sur les deux
    # premières lignes de la parution, et rien d'autre.
    apsis = [ligne for ligne in nommes if ligne[1] == "301294831" and ligne[3] == "101"]
    assert len(apsis) == 1
    assert apsis[0][0] == "57672"
    assert apsis[0][2] == "APSIS EMERGENCE"
    assert apsis[0][5] == 37_180.0


def test_les_beneficiaires_se_classent_du_plus_gros_au_plus_petit(lignes):
    nommes = s.beneficiaires(lignes)
    par_commune = defaultdict(list)
    for code, _, _, _, _, montant in nommes:
        par_commune[code].append(montant)
    for code, montants in par_commune.items():
        assert montants == sorted(montants, reverse=True), f"{code} n'est pas classée"


def test_la_liste_nominative_exclut_exactement_ce_que_l_agregat_exclut(lignes):
    """Deux listes qui ne filtreraient pas pareil donneraient deux totaux.

    L'étranger et l'outre-mer hors départements ne sont rattachés à aucune
    commune ; les arrondissements municipaux se replient sur leur commune. La
    liste nominative applique les mêmes règles que l'agrégat publié juste
    au-dessus d'elle.
    """
    nommes = s.beneficiaires(lignes)
    communes = {code for code, *_ in nommes}
    assert not [code for code in communes if code.startswith(("99", "98", "975", "977", "978"))]
    # Paris, Lyon et Marseille, jamais leurs arrondissements.
    assert "75056" in communes
    assert not [code for code in communes if code.startswith(("751", "6938", "132"))]


def test_un_versement_sans_siren_reste_dans_l_agregat_mais_ne_nomme_personne(lignes):
    """Sans identifiant, on ne peut ni regrouper ni vérifier : on ne nomme pas.

    Les 660 lignes que le jaune marque « N/A » comptent dans le montant
    communal — c'est de l'argent versé — mais n'apparaissent sous aucun nom.
    """
    sans = [ligne for ligne in lignes if ligne["siren"] in ("", s.SIREN_ABSENT)]
    assert sans, "la parution porte des bénéficiaires non identifiés"
    nommes = s.beneficiaires(lignes)
    assert not [ligne for ligne in nommes if ligne[1] in ("", s.SIREN_ABSENT)]


def test_l_objet_retenu_est_celui_du_plus_gros_versement(lignes):
    """Une association reçoit au titre de dizaines d'objets différents.

    Les concaténer ferait un paragraphe illisible sous chaque nom ; en prendre
    un au hasard ferait dépendre l'affichage de l'ordre du fichier. C'est celui
    du versement le plus important qui représente le couple.
    """
    from plateforme.normalize.geo import commune_mere

    nommes = {(code, siren, programme): objet
              for code, siren, _, programme, objet, _ in s.beneficiaires(lignes)}
    # Un même SIREN peut avoir des établissements dans deux communes : la clé
    # les sépare, et les candidats doivent l'être aussi.
    versements = defaultdict(list)
    for ligne in lignes:
        if not ligne["objet"] or not ligne["siren"]:
            continue
        code = commune_mere(ligne["commune"]) or ligne["commune"]
        versements[(code, ligne["siren"], ligne["programme"])].append(ligne)
    verifies = 0
    for cle, objet in nommes.items():
        candidats = versements.get(cle)
        if objet is None or not candidats:
            continue
        assert objet == max(candidats, key=lambda ligne: ligne["montant"])["objet"]
        verifies += 1
    assert verifies > 10_000, "le contrôle doit porter sur la parution, pas sur trois lignes"


def test_l_ecriture_nominative_contre_un_vrai_entrepot(entrepot_seme, lignes):
    from plateforme import entrepot

    _semer_le_jeu(entrepot_seme)
    s.declarer(entrepot_seme)
    run_id = entrepot.start_run(entrepot_seme, s.DATASET, "manual")
    nommes = [ligne for ligne in s.beneficiaires(lignes) if ligne[0] == "57672"]
    ecrites = s.ecrire_beneficiaires(entrepot_seme, run_id, nommes)
    entrepot_seme.commit()
    assert ecrites == len(nommes)
    (compte, masse) = entrepot_seme.execute(
        "select count(*), sum(amount) from fin.public_subsidies where geo_code = '57672'"
    ).fetchone()
    assert compte == len(nommes)
    assert round(masse, 2) == round(math.fsum(ligne[5] for ligne in nommes), 2)
    # Recharger le même exercice remplace, il n'empile pas : sans le `delete`,
    # une seconde parution doublerait chaque bénéficiaire.
    s.ecrire_beneficiaires(entrepot_seme, run_id, nommes)
    entrepot_seme.commit()
    (apres,) = entrepot_seme.execute(
        "select count(*) from fin.public_subsidies where geo_code = '57672'"
    ).fetchone()
    assert apres == compte


def test_les_beneficiaires_se_rassemblent_par_programme(entrepot_seme):
    """Le simulateur règle un budget national : il lui faut les bénéficiaires
    par programme, pas cent un fichiers départementaux.

    Et surtout : ce fichier ne décompose rien. Il publie le total **déclaré**
    et le nombre de bénéficiaires, jamais une part du programme — les
    subventions nominatives ne redonnent pas les crédits dont elles relèvent."""
    from plateforme import entrepot, publish

    _semer_le_jeu(entrepot_seme)
    run_id = entrepot.start_run(entrepot_seme, s.DATASET, "manual")
    for code, siren, nom, programme, objet, montant in (
        ("33063", "111", "GRANDE ASSO", "163", "Fonctionnement", 500_000.0),
        ("75056", "222", "AUTRE ASSO", "163", "Innovation", 200_000.0),
        ("75056", "333", "PETITE ASSO", "163", "Fonctionnement", 1_000.0),
        ("33063", "444", "ASSO SPORT", "219", "Sport", 90_000.0),
    ):
        entrepot_seme.execute(
            "insert into fin.public_subsidies (fiscal_year, geo_level, geo_code,"
            " geo_vintage, beneficiary_siren, beneficiary_name, programme, purpose,"
            " amount, dataset_id, run_id) values (2023, 'commune', ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (code, s.MILLESIME, siren, nom, programme, objet, montant, s.DATASET, run_id),
        )
    entrepot_seme.commit()

    publie = publish.subventions_par_programme(entrepot_seme, plafond=2)
    assert publie["exercice"] == "2023"
    p163 = publie["programmes"]["163"]
    # Le total déclaré somme les trois versements, les deux communes confondues.
    assert p163["declare"] == 701_000.0
    # Le compte dit ce que la liste bornée ne montre pas.
    assert p163["beneficiaires_total"] == 3
    assert [b["nom"] for b in p163["beneficiaires"]] == ["GRANDE ASSO", "AUTRE ASSO"]
    # Aucune part du programme : c'est le chiffre qui ferait mentir la page.
    assert "part_du_programme" not in p163
    assert set(publie["programmes"]) == {"163", "219"}
