"""Subventions de l'État aux associations, commune par commune (jaune budgétaire).

10,54 milliards d'euros versés en 2021 à 55 500 établissements associatifs, une
ligne par versement : c'est le seul document public qui descende du budget de
l'État jusqu'au bénéficiaire nommé. Le jaune annexé au PLF 2023 en donne
102 619 lignes, réparties sur 110 programmes budgétaires.

**L'exercice est 2021, et il faut le dire.** C'est la dernière parution du jaune
qui liste les crédits attribués bénéficiaire par bénéficiaire. Cinq ans séparent
ce versement du lecteur ; la fiche le porte, faute de quoi ce montant serait lu
comme celui de l'année en cours.

**La commune est celle de l'établissement, pas celle de l'action.** Une
fédération nationale domiciliée dans le 15e arrondissement apparaît à Paris,
qu'elle intervienne à Guéret ou à Mayotte. C'est pourquoi Paris pèse 3,23 Md€,
soit 32,5 % du total publié, et les dix premières communes 43,5 % : cette carte
dit où sont les sièges, et l'écrire est la moitié du travail.

**Sirene est le témoin, pas la source. C'est la décision d'architecture de ce
module.** Le fichier porte déjà, pour chacune de ses lignes, un code commune
(colonne « COG : code », servie à 100 %) issu de l'appariement Sirene qu'a fait
l'administration au moment de la parution. Ce code est *contemporain du
versement* ; le répertoire Sirene d'aujourd'hui, lui, donne l'adresse
d'aujourd'hui. Publier celle-ci reviendrait à déménager rétroactivement les
associations qui ont bougé depuis. Le code du jaune est donc celui qui est
publié, et Sirene sert à le **vérifier** : deux codes commune produits
indépendamment, à cinq ans d'écart, pour le même SIRET. Ils s'accordent sur
99,97 % des euros.

Charger tout Sirene dans l'entrepôt aurait été le troisième chemin possible.
C'est un jeu à part entière — 43,9 millions d'établissements, 54 colonnes,
2,2 Go — qui ne serait ici qu'un fichier de correspondance. Les 55 500 SIRET du
jaune se résolvent en 45 secondes par une requête DuckDB sur le parquet
*distant*, sans rien rapatrier ; c'est ce que fait `resoudre()`, seul appel
réseau à Sirene du module, et les tests ne l'appellent pas.

**Ce que Sirene apporte que le jaune n'a pas.** La colonne « Etat administratif »
du fichier vaut « Non déterminé » sur ses 102 619 lignes : elle ne dit rien.
Sirene dit que 21 181 lignes — 20,8 % — viennent d'établissements aujourd'hui
**fermés**, pour 1,55 Md€. La subvention a bien été versée en 2021 ;
l'établissement n'existe plus. C'est une information à publier, pas à taire.

**Ce qui n'est pas une commune française est écarté, jamais réparti.** La
colonne géographique porte aussi des codes pays (99xxx) et des collectivités du
Pacifique et de l'Atlantique (98xxx, 975, 977, 978). Cela pèse 600,8 M€, dont
527,1 M€ pour la seule Association internationale de développement, à
Washington : c'est la plus grosse ligne du fichier, et la diluer sur les
communes françaises fabriquerait une subvention que personne n'a reçue.

**Le contrôle qui bloque** compare deux mondes plutôt qu'une source à
elle-même (docs/couverture.py) : pour chaque ligne portant un SIRET, le code
commune du jaune doit être celui que Sirene donne pour ce SIRET. Une colonne qui
aurait glissé, un SIRET reconstruit de travers, un zéro de tête perdu à la
lecture du tableur : rien de tout cela ne peut satisfaire cette égalité par
hasard, là où un total resterait parfaitement égal à lui-même.

⚠ Les montants sont écrits par le tableur avec des **espaces insécables** en
séparateur de milliers et une **virgule** décimale (`   3 850,00`). Un
`float()` naïf lève, un `replace(" ", "")` sur l'espace ordinaire ne fait rien :
c'est `_nombre()` qui traite les deux.

Usage : python -m plateforme.normalize.subventions [--store r2:plateforme-raw]
"""

import argparse
import io
import json
import math
import unicodedata
from collections import defaultdict

from plateforme import couverture, entrepot, revisions
from plateforme.http import fetch, telecharger
from plateforme.limites import garde_fou_volume
from plateforme.normalize.geo import MILLESIME, commune_mere, make_store
from plateforme.normalize.ofgl import filtrer_territoires_connus

DATASET = "plf-2023-effort-associations"
SOURCE = "data-economie"
JEU = (
    "plf-2023-effort-financier-de-l-etat-en-faveur-des-associations"
    "-liste-des-credits"
)
PIECE = (
    "plf_2023_effort_financier_de_l_etat_en_faveur_des_associations"
    "_liste_des_credits_attribues_xlsx"
)
URL = f"https://data.economie.gouv.fr/api/v2/catalog/datasets/{JEU}/attachments/{PIECE}"
FEUILLE = "versements 2021"

# L'exercice des versements, et non celui du projet de loi de finances qui les
# annexe. Le jaune est celui du PLF 2023 ; l'argent est parti en 2021.
EXERCICE = "2021"

# Sirene : le jeu data.gouv et le titre de la ressource. L'URL du parquet porte
# la date de production (…/20260801-074451/…) et change à chaque parution
# mensuelle ; la coder en dur reviendrait à vérifier éternellement contre un
# répertoire périmé. Elle est donc retrouvée par le catalogue, comme
# `fonds_europeens` retrouve son classeur sur la page de l'ANCT.
SIRENE_JEU = "base-sirene-des-entreprises-et-de-leurs-etablissements-siren-siret"
SIRENE_RESSOURCE = "StockEtablissement"
SIRENE_CATALOGUE = f"https://www.data.gouv.fr/api/1/datasets/{SIRENE_JEU}/"

# Les sept colonnes lues sur les dix-sept du fichier.
#
# La dénomination et l'objet ont d'abord été refusés : un agrégat communal n'a
# pas besoin de savoir qui a reçu. Mais « 353 € par habitant » ne répond pas à
# la question qu'on vient poser — *à quelles associations* — et un total sans
# ses bénéficiaires demande de croire sur parole. Ce document existe pour être
# lu bénéficiaire par bénéficiaire ; le lire à moitié était le trahir.
#
# Les autres colonnes restent refusées, et les nommer vaut refus explicite :
# convention, date de création, catégorie juridique, activité et sa
# nomenclature, économie sociale et solidaire, libellé de commune (redondant
# avec le code). « Siège » et « Etat administratif » ne sont pas lues parce
# qu'elles ne disent rien : la première vaut « Non » et la seconde « Non
# déterminé » sur les 102 619 lignes. C'est Sirene qui donne l'état
# administratif.
COLONNES_LUES = {
    "programme": "Programme",
    "siren": "SIREN",
    "nic": "NIC",
    "montant": "Montant",
    "commune": "COG : code",
    "nom": "Dénomination",
    "objet": "Objet 2021",
}

# Le décompte de la parution chargée. Ce n'est pas une invariante interne : la
# somme de ce qu'on a lu est toujours égale à elle-même, et un classeur lu à
# moitié rend un total parfaitement cohérent. Le relever demande d'avoir regardé
# la nouvelle parution.
LIGNES = 102_619

# La valeur par laquelle le jaune marque un bénéficiaire qu'il n'a pas su
# rattacher au répertoire : 660 lignes, presque toutes hors de France.
SIREN_ABSENT = "N/A"

# Codes de la colonne géographique qui ne désignent aucune commune du Code
# officiel géographique. 99xxx est la liste des pays de l'INSEE ; 98xxx, 975,
# 977 et 978 sont les collectivités d'outre-mer, qui ont des communes mais pas
# dans `v_commune_<millésime>.csv`. Les écarter nommément, plutôt que les
# laisser tomber au filtre du référentiel, est ce qui permet à la couverture de
# ne mesurer que les vrais ratés d'appariement.
PAYS = "99"
OUTRE_MER = ("98", "975", "977", "978")

# Seuils du contrôle bloquant, mesurés sur la parution du PLF 2023 confrontée au
# stock Sirene du 1er août 2026 : 99,967 % des euros portant un SIRET se
# résolvent, et 99,974 % de ceux-là tombent sur la même commune des deux côtés.
# Un pour cent de marge absorbe cinq ans de déménagements et de radiations ; une
# colonne décalée ferait chuter la concordance à quelques pour cent.
RESOLUTION_MINIMALE = 0.99
CONCORDANCE_MINIMALE = 0.99

THEME = "vie_associative"

INDICATEURS = {
    "etat_subventions_associations": (
        "Subventions de l'État aux associations",
        "Ce que l'État a versé aux associations en 2021, dans la commune de"
        " l'établissement bénéficiaire, pas celle où l'action est menée. Une"
        " fédération nationale compte là où elle est domiciliée : Paris pèse un"
        " tiers du total. Ce sont des versements de l'État, pas des"
        " collectivités.",
        "EUR",
        "Somme des versements de l'État aux établissements associatifs domiciliés"
        " dans la commune, exercice 2021",
    ),
    "etat_subventions_associations_etablissements": (
        "Établissements associatifs subventionnés par l'État",
        "Combien d'établissements associatifs domiciliés dans la commune ont reçu"
        " au moins un versement de l'État en 2021. Ce nombre sert de dénominateur"
        " au montant : un million d'euros pour une association et pour trois"
        " cents ne se lisent pas de la même façon.",
        "count",
        "Nombre de SIRET distincts ayant reçu au moins un versement, exercice 2021",
    ),
}

TECHNIQUE = (
    "Versements de l'État aux associations au titre de l'exercice **2021**,"
    " source : jaune budgétaire « Effort financier de l'État en faveur des"
    " associations » annexé au projet de loi de finances pour 2023, liste des"
    " crédits attribués. Le fichier donne 102 619 versements pour"
    " 10,54 Md€ sur 110 programmes budgétaires. **L'exercice a cinq ans** :"
    " c'est la dernière parution du jaune qui descende au bénéficiaire nommé, et"
    " ce montant ne dit rien de l'année en cours. **Ce sont des versements de"
    " l'État**, sur son budget : les subventions des communes, départements et"
    " régions aux associations n'y sont pas. Elles se lisent dans les comptes"
    " des collectivités (OFGL, poste « subventions aux personnes de droit"
    " privé »). **La commune est celle de l'établissement bénéficiaire, pas"
    " celle où l'action est menée** : une fédération nationale est comptée à"
    " l'adresse de son siège, si bien que Paris porte 3,23 Md€, soit 32,5 % du"
    " total publié, et les dix premières communes 43,5 %. Le code commune est"
    " celui que le jaune porte lui-même (colonne « COG : code »), contemporain"
    " du versement ; il est vérifié ligne à ligne contre le code que le"
    " répertoire Sirene donne aujourd'hui pour le même SIRET, et les deux"
    " s'accordent sur 99,97 % des euros. Sirene indique par ailleurs que"
    " 21 181 versements (20,8 % des lignes, 1,55 Md€) sont allés à des"
    " établissements **aujourd'hui fermés** : la subvention a bien été versée en"
    " 2021, l'établissement n'existe plus. Les arrondissements municipaux de"
    " Paris, Lyon et Marseille sont repliés sur leur commune. Ne sont pas"
    " publiés, et ne sont répartis sur aucune commune : les bénéficiaires situés"
    " à l'étranger (164 lignes, 569,1 M€, dont 527,1 M€ pour la seule"
    " Association internationale de développement aux États-Unis, plus grosse"
    " ligne du fichier), ceux des collectivités d'outre-mer hors départements"
    " (727 lignes, 31,7 M€) et 13 codes communaux disparus du Code officiel"
    " géographique depuis 2021 (1,76 M€). Quatre lignes portent un montant"
    " négatif (-22 674 € au total) : ce sont des reversements. Le dénombrement"
    " des établissements compte les SIRET distincts ; 24 versements métropolitains"
    " ou ultramarins n'en portent aucun et n'y entrent donc pas, alors qu'ils"
    " comptent dans le montant."
)


class ConcordanceRompue(RuntimeError):
    """Le code commune du jaune et celui de Sirene ne décrivent plus le même monde."""


def _nombre(valeur) -> float | None:
    """Un montant du classeur -> float, ou None s'il n'est pas renseigné.

    Le tableur écrit `\xa0\xa0\xa0\xa0\xa03\xa0850,00` : espaces insécables en
    rembourrage *et* en séparateur de milliers, virgule décimale. Les trois
    pièges sont traités ensemble ; ne traiter que l'espace ordinaire laisse
    passer une chaîne que `float()` refuse, et l'échec est alors silencieux si
    l'appelant l'attrape.
    """
    if valeur is None:
        return None
    if isinstance(valeur, (int, float)):
        return float(valeur)
    texte = str(valeur).replace("\xa0", " ").replace(" ", "").replace(",", ".")
    return float(texte) if texte else None


def _commune(valeur) -> str:
    """Le code de la colonne « COG : code », rendu à cinq caractères.

    Le tableur type la colonne en nombre partout où elle en est un : l'Ain rend
    1004 pour 01004, et la Corse reste en texte (2A004). Sans le rembourrage,
    5 248 lignes tomberaient hors référentiel sans qu'aucun total ne bouge.
    """
    texte = str(valeur).strip() if valeur is not None else ""
    return texte.zfill(5) if texte.isdigit() else texte


def _siret(siren, nic) -> str:
    """SIREN + NIC -> SIRET à quatorze chiffres, ou "" si le jaune n'en a pas.

    Les deux colonnes sont typées en nombre par le tableur, qui mange les zéros
    de tête : le NIC 00088 revient en 88, et 24 SIREN commençant par zéro
    reviennent à huit chiffres. Le rembourrage est ce qui rend la clé à Sirene.
    """
    if siren is None or nic is None or str(siren).strip() == SIREN_ABSENT:
        return ""
    return f"{int(siren):09d}{int(nic):05d}"


def _texte(valeur) -> str:
    """Le tableur sépare tous ses mots par des espaces insécables.

    « APSIS\xa0EMERGENCE » et « Transfert\xa0direct\xa0aux\xa0ménages » : le
    document entier est composé ainsi. Publié tel quel, le nom se recopie mal,
    ne se cherche pas et coupe au mauvais endroit. `NFKC` ramène ces espaces à
    l'espace ordinaire, et avec elles les autres caractères de compatibilité.
    """
    if valeur is None:
        return ""
    return " ".join(unicodedata.normalize("NFKC", str(valeur)).split())


def lire(contenu: bytes) -> list[dict]:
    """Le classeur du jaune -> [{programme, siret, montant, commune}].

    Les repères sont contrôlés plutôt que supposés : la feuille est nommée, et
    les cinq colonnes lues sont retrouvées par leur intitulé exact. Les deux
    premières rangées portent le titre du tableau et une ligne vide ; l'en-tête
    est cherché plutôt que compté, un tableur en gagnant une à chaque parution.
    """
    import openpyxl

    livre = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    if FEUILLE not in livre.sheetnames:
        raise ValueError(
            f"feuille « {FEUILLE} » absente du classeur {livre.sheetnames} :"
            " le jaune a changé de forme ou d'exercice"
        )
    rangees = livre[FEUILLE].iter_rows(values_only=True)
    entete = None
    for rangee in rangees:
        valeurs = {str(c).strip(): rang for rang, c in enumerate(rangee) if c is not None}
        if all(nom in valeurs for nom in COLONNES_LUES.values()):
            entete = valeurs
            break
    if entete is None:
        raise ValueError(
            f"colonnes {sorted(COLONNES_LUES.values())} introuvables dans"
            f" « {FEUILLE} » : le jaune a changé de colonnes"
        )

    lignes = []
    for rangee in rangees:
        brut = {cle: rangee[entete[nom]] for cle, nom in COLONNES_LUES.items()}
        montant = _nombre(brut["montant"])
        if montant is None:
            continue
        lignes.append({
            "programme": _texte(brut["programme"]),
            "siren": _texte(brut["siren"]),
            "siret": _siret(brut["siren"], brut["nic"]),
            "montant": round(montant, 2),
            "commune": _commune(brut["commune"]),
            "nom": _texte(brut["nom"]),
            "objet": _texte(brut["objet"]),
        })
    return lignes


def url_sirene(catalogue: str) -> str:
    """L'URL du StockEtablissement en parquet, retrouvée dans le catalogue.

    L'URL porte la date de production et change chaque mois. La chercher plutôt
    que la coder en dur évite de vérifier éternellement le jaune contre un
    répertoire périmé — et de ne pas s'en apercevoir, la concordance restant
    bonne pendant des années.
    """
    ressources = json.loads(catalogue)["resources"]
    parquets = [
        r["url"]
        for r in ressources
        if SIRENE_RESSOURCE in (r.get("title") or "")
        and (r.get("format") or "").lower() == "parquet"
        and "Historique" not in (r.get("title") or "")
        and "LiensSuccession" not in (r.get("title") or "")
    ]
    if len(parquets) != 1:
        raise ValueError(
            f"une seule ressource « {SIRENE_RESSOURCE} … (format parquet) » est"
            f" attendue dans {SIRENE_JEU}, {len(parquets)} trouvée(s) : le"
            " catalogue de data.gouv a changé, il faut regarder avant de recharger"
        )
    return parquets[0]


def resoudre(adresse: str, sirets: list[str]) -> dict[str, tuple[str, str]]:
    """{SIRET: (code commune, état administratif)} — **le seul appel réseau à Sirene**.

    Le fichier fait 2,2 Go et 43,9 millions de lignes ; il n'est pas
    téléchargé. DuckDB lit le parquet à distance et ne rapatrie que les groupes
    de lignes utiles : 55 500 SIRET se résolvent en une quarantaine de secondes.
    C'est ce qui permet à Sirene d'être un témoin plutôt qu'un jeu de plus dans
    l'entrepôt.

    La connexion est neuve et en mémoire, jamais celle de l'entrepôt : ni
    l'extension `httpfs` ni la table de travail n'ont à toucher la base de
    production. Les tests ne passent pas par ici — ils reçoivent la
    correspondance déjà résolue.
    """
    import duckdb

    conn = duckdb.connect()
    try:
        conn.execute("install httpfs; load httpfs;")
        conn.execute(
            "create temp table cibles as select unnest($sirets) as siret",
            {"sirets": list(sirets)},
        )
        lignes = conn.execute(
            "select e.siret, e.codeCommuneEtablissement,"
            " e.etatAdministratifEtablissement"
            f" from read_parquet('{adresse}') e join cibles c on e.siret = c.siret"
        ).fetchall()
    finally:
        conn.close()
    return {siret: (commune or "", etat or "") for siret, commune, etat in lignes}


def controler(lignes: list[dict], sirene: dict[str, tuple[str, str]]) -> dict:
    """Le contrôle qui bloque : le code commune du jaune se retrouve dans Sirene.

    Deux codes commune produits indépendamment, à cinq ans d'écart, pour le même
    établissement — l'un par l'administration au moment de la parution du jaune,
    l'autre par le répertoire d'aujourd'hui. Aucun total interne ne peut voir ce
    que celui-ci voit : une colonne décalée, un SIRET reconstruit sans ses zéros
    de tête ou un code commune tronqué laissent la somme du fichier rigoureusement
    inchangée et font tomber la concordance à quelques pour cent.

    Les deux taux sont mesurés **en euros** et non en lignes : mille versements
    de cent euros ne pèsent pas ce que pèse la ligne de 527 M€.
    """
    if not lignes:
        raise ConcordanceRompue("le classeur est vide")
    if len(lignes) != LIGNES:
        raise ConcordanceRompue(
            f"{len(lignes)} versements lus contre {LIGNES} attendus. Lecture"
            " tronquée, ou nouvelle parution : dans le second cas, il faut avoir"
            " regardé le jaune avant de relever le décompte attendu."
        )
    identifiees = [ligne for ligne in lignes if ligne["siret"]]
    if not identifiees:
        raise ConcordanceRompue(
            "aucune ligne ne porte de SIRET : la clé vers Sirene est perdue et le"
            " contrôle ne vérifie plus rien"
        )
    resolues = [
        ligne for ligne in identifiees if sirene.get(ligne["siret"], ("", ""))[0]
    ]
    masse_identifiees = math.fsum(ligne["montant"] for ligne in identifiees)
    masse_resolues = math.fsum(ligne["montant"] for ligne in resolues)
    resolution = masse_resolues / masse_identifiees if masse_identifiees else 0.0
    if resolution < RESOLUTION_MINIMALE:
        raise ConcordanceRompue(
            f"{len(resolues)} versements sur {len(identifiees)} portent un SIRET"
            f" que Sirene connaît et localise, soit {100 * resolution:.2f} % des"
            " euros identifiés. Les SIRET du jaune ne sont pas ceux du répertoire :"
            " zéros de tête perdus à la lecture du tableur, ou colonnes SIREN et"
            " NIC inversées."
        )

    divergents = [
        ligne for ligne in resolues if sirene[ligne["siret"]][0] != ligne["commune"]
    ]
    masse_divergents = math.fsum(ligne["montant"] for ligne in divergents)
    concordance = (
        (masse_resolues - masse_divergents) / masse_resolues if masse_resolues else 0.0
    )
    if concordance < CONCORDANCE_MINIMALE:
        divergent = divergents[0]
        raise ConcordanceRompue(
            f"{len(divergents)} versements sur {len(resolues)} placent"
            f" l'établissement ailleurs que Sirene, soit {100 * (1 - concordance):.2f} %"
            f" des euros résolus — par exemple le SIRET {divergent['siret']},"
            f" commune {divergent['commune']} dans le jaune et"
            f" {sirene[divergent['siret']][0]} dans Sirene. Le code commune publié"
            " ne décrit plus l'établissement qui a reçu l'argent."
        )

    fermes = [ligne for ligne in resolues if sirene[ligne["siret"]][1] == "F"]
    return {
        "versements": len(lignes),
        "masse_totale": round(math.fsum(ligne["montant"] for ligne in lignes), 2),
        "programmes": len({ligne["programme"] for ligne in lignes}),
        "etablissements": len({ligne["siret"] for ligne in identifiees}),
        "sans_siret": len(lignes) - len(identifiees),
        "taux_resolution_euros": resolution,
        "taux_concordance_euros": concordance,
        "desaccords": len(divergents),
        "masse_desaccords": round(masse_divergents, 2),
        # Ce que Sirene apprend et que le jaune ne dit pas : sa colonne
        # « Etat administratif » vaut « Non déterminé » sur toutes ses lignes.
        "versements_etablissement_ferme": len(fermes),
        "masse_etablissement_ferme": round(
            math.fsum(ligne["montant"] for ligne in fermes), 2
        ),
    }


def hors_territoire(code: str) -> str | None:
    """Le motif pour lequel un code n'est pas une commune, ou None.

    Écarter nommément vaut mieux que laisser tomber au filtre du référentiel :
    ces codes sont des territoires réels, simplement absents du fichier communal
    du Code officiel géographique. Les confondre avec des ratés d'appariement
    ferait passer la couverture de 99,98 % à 94,3 % et masquerait, sous le
    dépassement de seuil, les treize communes réellement perdues.
    """
    if code.startswith(PAYS):
        return "pays_etranger"
    if code.startswith(OUTRE_MER):
        return "collectivite_outre_mer"
    return None


def par_commune(lignes: list[dict]) -> tuple[list[tuple], dict[str, dict]]:
    """-> (observations candidates, ce qui est écarté par motif).

    Les arrondissements municipaux de Paris, Lyon et Marseille se replient sur
    leur commune : le jaune ne code que par arrondissement, si bien que laisser
    75108 tel quel publierait Paris à zéro — c'est exactement le défaut que
    `couverture.py` raconte pour le répertoire des logements sociaux.
    """
    montants: dict[str, list[float]] = defaultdict(list)
    etablissements: dict[str, set[str]] = defaultdict(set)
    ecartes: dict[str, list[float]] = defaultdict(list)
    for ligne in lignes:
        motif = hors_territoire(ligne["commune"])
        if motif:
            ecartes[motif].append(ligne["montant"])
            continue
        code = commune_mere(ligne["commune"]) or ligne["commune"]
        montants[code].append(ligne["montant"])
        if ligne["siret"]:
            etablissements[code].add(ligne["siret"])

    candidates = [
        ("etat_subventions_associations", "commune", code, EXERCICE,
         round(math.fsum(valeurs), 2))
        for code, valeurs in sorted(montants.items())
    ]
    candidates += [
        ("etat_subventions_associations_etablissements", "commune", code, EXERCICE,
         float(len(etablissements[code])))
        for code in sorted(montants)
    ]
    resume = {
        motif: {"versements": len(valeurs), "montant": round(math.fsum(valeurs), 2)}
        for motif, valeurs in sorted(ecartes.items())
    }
    return candidates, resume


def masse_par_maille(candidates) -> dict[str, float]:
    """Les euros par maille, pour les deux côtés de la couverture.

    Seul l'indicateur monétaire y entre : mélanger des euros et des comptes
    d'établissements ferait un dénominateur que rien ne mesure.
    """
    par_maille: dict[str, list[float]] = defaultdict(list)
    for identifiant, niveau, _, _, valeur in candidates:
        if identifiant == "etat_subventions_associations":
            par_maille[niveau].append(valeur)
    return {
        niveau: round(math.fsum(valeurs), 2)
        for niveau, valeurs in par_maille.items()
    }


def declarer(conn) -> None:
    with conn.cursor() as curseur:
        for identifiant, (libelle, publique, unite, formule) in sorted(INDICATEURS.items()):
            definition = curseur.execute(
                """
                insert into core.indicator_definitions
                    (public_definition, technical_definition, formula, unit_notes,
                     confidence_level, badges)
                values (?, ?, ?, ?, 'observed', array['Officiel','Donnée brute'])
                returning definition_id
                """,
                (
                    publique,
                    TECHNIQUE,
                    formule,
                    "euros courants versés en 2021"
                    if unite == "EUR"
                    else "établissements (SIRET) distincts",
                ),
            ).fetchone()[0]
            # Un dénombrement d'établissements n'est ni en euros courants ni dans
            # une comptabilité budgétaire : les deux colonnes restent nulles
            # plutôt que de recopier celles du montant.
            monetaire = unite == "EUR"
            curseur.execute(
                """
                insert into core.indicators
                    (indicator_id, dataset_id, definition_id, theme, label_fr, unit,
                     additive, price_basis, accounting_frame, geo_levels,
                     time_granularity, published)
                values (?, ?, ?, ?, ?, ?, true, ?, ?, array['commune'],
                        'annuelle', true)
                on conflict (indicator_id) do update set unit = excluded.unit,
                    definition_id = excluded.definition_id,
                    label_fr = excluded.label_fr,
                    theme = excluded.theme, published = true
                """,
                (
                    identifiant, DATASET, definition, THEME, libelle, unite,
                    "current" if monetaire else None,
                    "budgetaire" if monetaire else None,
                ),
            )
        curseur.execute(
            "delete from core.indicator_definitions d where not exists"
            " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
        )
    conn.commit()


COLONNES = ("value",)


def beneficiaires(lignes: list[dict]) -> list[tuple]:
    """-> [(commune, SIREN, nom, programme, objet, montant)], la plus grosse d'abord.

    Le grain est le couple (bénéficiaire, programme) : le jaune porte une ligne
    par versement, et une même association en reçoit plusieurs au titre d'un
    même programme. Les sommer ici plutôt qu'à la lecture évite qu'un total
    dépende de qui l'additionne.

    Les mêmes exclusions que l'agrégat, et pour les mêmes raisons : les
    bénéficiaires hors du territoire ne sont rattachés à aucune commune, et les
    arrondissements municipaux se replient sur leur commune. Deux listes qui ne
    filtreraient pas pareil donneraient un total de bénéficiaires différent du
    total publié juste au-dessus.

    Un versement sans SIREN identifiable n'est pas publié nommément : le jaune
    marque 660 lignes ainsi, et sans identifiant on ne peut ni les regrouper ni
    les vérifier. Ils restent dans l'agrégat communal, qui ne nomme personne.
    """
    montants: dict[tuple, float] = defaultdict(float)
    noms: dict[tuple, str] = {}
    objets: dict[tuple, tuple[float, str]] = {}
    for ligne in lignes:
        if hors_territoire(ligne["commune"]) or not ligne["siren"] or ligne["siren"] == SIREN_ABSENT:
            continue
        code = commune_mere(ligne["commune"]) or ligne["commune"]
        cle = (code, ligne["siren"], ligne["programme"])
        montants[cle] += ligne["montant"]
        noms.setdefault(cle, ligne["nom"])
        # L'objet du plus gros versement représente le couple : les objets se
        # comptent par dizaines pour une association, et les concaténer ferait
        # un paragraphe illisible sous chaque nom.
        if ligne["objet"] and ligne["montant"] > objets.get(cle, (0.0, ""))[0]:
            objets[cle] = (ligne["montant"], ligne["objet"])
    return sorted(
        (
            (code, siren, noms[(code, siren, programme)], programme,
             objets.get((code, siren, programme), (0.0, ""))[1] or None,
             round(montant, 2))
            for (code, siren, programme), montant in montants.items()
        ),
        key=lambda ligne: (ligne[0], -ligne[5], ligne[1]),
    )


def ecrire_beneficiaires(conn, run_id, lignes: list[tuple]) -> int:
    """Remplace la liste nominative de l'exercice.

    Un `delete` puis un `insert`, et non `revisions.remplacer` : ces lignes ne
    sont pas des observations d'indicateur — elles n'ont ni identifiant
    d'indicateur ni série dans le temps, et rien à archiver. Le jaune ne
    republie pas un exercice passé : à chaque parution son exercice.
    """
    conn.execute("delete from fin.public_subsidies where fiscal_year = ?", (EXERCICE,))
    conn.executemany(
        """
        insert into fin.public_subsidies
            (fiscal_year, geo_level, geo_code, geo_vintage, beneficiary_siren,
             beneficiary_name, programme, purpose, amount, dataset_id, run_id)
        values (?, 'commune', ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (int(EXERCICE), code, MILLESIME, siren, nom, programme, objet, montant,
             DATASET, run_id)
            for code, siren, nom, programme, objet, montant in lignes
        ],
    )
    return len(lignes)


def run(store_spec: str) -> int:
    conn = entrepot.connect()
    store = make_store(store_spec)
    declarer(conn)
    run_id = entrepot.start_run(conn, DATASET, "manual")
    try:
        garde_fou_volume(conn)
        contenu = telecharger(URL, timeout=600)
        entrepot.record_asset(
            conn, store, run_id, DATASET, SOURCE,
            "effort-financier-associations-2021.xlsx", contenu, URL,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        lignes = lire(contenu)
        entrepot.etape(f"{len(lignes)} versements lus, exercice {EXERCICE}")

        adresse = url_sirene(fetch(SIRENE_CATALOGUE, timeout=120).text)
        sirets = sorted({ligne["siret"] for ligne in lignes if ligne["siret"]})
        entrepot.etape(f"{len(sirets)} SIRET à résoudre sur {adresse}")
        sirene = resoudre(adresse, sirets)
        verifies = controler(lignes, sirene)
        entrepot.etape(
            f"{100 * verifies['taux_concordance_euros']:.3f} % des euros résolus"
            " tombent sur la même commune dans le jaune et dans Sirene"
        )

        candidates, ecartes_hors = par_commune(lignes)
        gardees, ecartes = filtrer_territoires_connus(conn, candidates)
        ecrites, revisees = revisions.remplacer(
            conn, run_id, sorted(INDICATEURS), COLONNES,
            [
                (identifiant, niveau, code, MILLESIME, periode, valeur)
                for identifiant, niveau, code, periode, valeur in gardees
            ],
        )
        nommes = ecrire_beneficiaires(conn, run_id, beneficiaires(lignes))
        entrepot.etape(f"{nommes} bénéficiaires nommés, exercice {EXERCICE}")

        lus = masse_par_maille(candidates)
        parts = couverture.controler(lus, masse_par_maille(gardees))
        conn.execute(
            """
            insert into meta.data_quality_checks
                (run_id, dataset_id, check_name, severity, passed, observed)
            values (?, ?, 'le_code_commune_du_jaune_se_retrouve_dans_sirene',
                    'blocker', true, ?)
            """,
            (run_id, DATASET, json.dumps({
                **verifies,
                "exercice": EXERCICE,
                "sirene": adresse,
                "hors_territoire": ecartes_hors,
                "masse_communale": lus.get("commune", 0.0),
                "hors_referentiel": sorted(ecartes),
                "couverture": {n: round(p, 6) for n, p in sorted(parts.items())},
            }, ensure_ascii=False)),
        )
        conn.commit()
        entrepot.finish_run(
            conn, run_id, "success", rows_read=len(lignes), rows_written=ecrites
        )
        print(
            f"subventions aux associations : {ecrites} observations, exercice"
            f" {EXERCICE}, {verifies['masse_totale'] / 1e9:.2f} Md€ versés dont"
            f" {lus.get('commune', 0.0) / 1e9:.2f} Md€ rattachés à une commune"
            f" française ; concordance avec Sirene"
            f" {100 * verifies['taux_concordance_euros']:.3f} % des euros,"
            f" {verifies['versements_etablissement_ferme']} versements à des"
            f" établissements fermés depuis ; {nommes} bénéficiaires nommés ;"
            f" {revisees} révisées, {len(ecartes)} hors référentiel"
        )
        return 0
    except Exception as error:  # noqa: BLE001 — tout échec finit tracé dans le lineage
        entrepot.finish_run(conn, run_id, "failed", error=str(error))
        raise
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default=".snapshots")
    return run(parser.parse_args().store)


if __name__ == "__main__":
    raise SystemExit(main())
