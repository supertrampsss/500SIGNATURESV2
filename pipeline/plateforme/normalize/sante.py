"""Accessibilité aux médecins généralistes (APL, DREES) vers core.observations.

« Y a-t-il un médecin près de chez moi ? » — la question des déserts médicaux.
L'APL (accessibilité potentielle localisée) de la DREES y répond mieux qu'une
densité : elle compte les **consultations réellement accessibles par an et par
habitant**, en tenant compte de la distance aux cabinets, de leur niveau
d'activité et des besoins de soins selon l'âge de la population locale. La
DREES retient 2,5 consultations/an/habitant comme seuil de sous-densité.

**Un indicateur modélisé, et la fiche le dit.** L'APL n'est pas un comptage
mais le résultat d'un modèle (flux domicile-cabinet à moins de 20 minutes,
activité standardisée) : deux millésimes se comparent, une décimale isolée ne
se surinterprète pas.

La source est un classeur Excel (une feuille par millésime) diffusé par la
DREES via data.gouv ; l'URL est relue dans le catalogue à chaque run, et deux
contrôles structurels bloquent si le fichier change de sens : l'en-tête doit
annoncer « Code commune INSEE », et la ligne d'unité doit contenir « par
habitant standardisé » — c'est elle qui garantit que le chiffre publié est
bien des consultations par habitant, pas autre chose.

Usage : python -m plateforme.normalize.sante [--store r2:plateforme-raw]
"""

import json
import argparse
import io
import re


from plateforme import entrepot, revisions
from plateforme.http import fetch
from plateforme.limites import garde_fou_volume
from plateforme.normalize.geo import MILLESIME, commune_mere, make_store
from plateforme.normalize.ofgl import filtrer_territoires_connus

DATASET = "drees-apl"
SOURCE = "data-gouv"
DATASET_DATAGOUV = "62263314072c63d4d53e0c50"
API_DATAGOUV = "https://www.data.gouv.fr/api/1"

INDICATEUR = "drees_apl_generalistes"
FICHE = {
    "libelle": "Accès aux médecins généralistes (APL)",
    "public": "Le nombre de consultations de médecin généraliste réellement"
    " accessibles par an et par habitant, selon la distance aux cabinets, leur"
    " activité et l'âge de la population. En dessous de 2,5, la DREES parle de"
    " sous-densité — ce qu'on appelle un désert médical.",
    "technique": "Accessibilité potentielle localisée (APL) aux médecins"
    " généralistes, DREES, en consultations/visites accessibles par an par"
    " habitant standardisé. Indicateur modélisé (accès à moins de 20 minutes,"
    " activité et besoins standardisés), pas un comptage.",
    "formule": "DREES, classeur APL, colonne « APL aux médecins généralistes »",
}

# Le plafond vise les artefacts de lecture — une colonne de population prise
# pour l'APL — pas les vraies valeurs hautes : un village alpin avec un
# médecin et peu d'habitants standardisés dépasse 40 (constaté : 05085 à
# 40,03 en 2023, d'abord écarté à tort par un plafond à 30).
PLAFOND_PLAUSIBLE = 100.0


def url_courante() -> str:
    """La dernière édition du classeur des généralistes, relue au catalogue."""
    catalogue = fetch(f"{API_DATAGOUV}/datasets/{DATASET_DATAGOUV}/", timeout=60).json()
    candidates = [
        r for r in catalogue["resources"]
        if "APL aux médecins généralistes" in r["title"]
    ]
    if not candidates:
        raise ValueError("classeur APL des généralistes absent du catalogue DREES")
    return max(candidates, key=lambda r: r.get("last_modified") or "")["url"]


def declarer(conn) -> None:
    with conn.cursor() as curseur:
        definition = curseur.execute(
            """
            insert into core.indicator_definitions
                (public_definition, technical_definition, formula, confidence_level, badges)
            values (?, ?, ?, 'computed', array['Officiel'])
            returning definition_id
            """,
            (FICHE["public"], FICHE["technique"], FICHE["formule"]),
        ).fetchone()[0]
        curseur.execute(
            """
            insert into core.indicators
                (indicator_id, dataset_id, definition_id, theme, label_fr, unit,
                 additive, geo_levels, time_granularity, published)
            values (?, ?, ?, 'sante', ?, 'consultations_par_an', false,
                    array['commune'], 'annuelle', true)
            on conflict (indicator_id) do update set
                definition_id = excluded.definition_id, label_fr = excluded.label_fr,
                theme = excluded.theme, published = true
            """,
            (INDICATEUR, DATASET, definition, FICHE["libelle"]),
        )
        curseur.execute(
            "delete from core.indicator_definitions d where not exists"
            " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
        )
    conn.commit()


def lire(contenu: bytes) -> tuple[list[tuple[str, str, float, float | None]], dict]:
    """Classeur APL -> [(code commune, millésime, valeur)], écartées.

    Chaque feuille « APL <année> » porte un bloc d'en-tête puis les données.
    Les repères structurels sont contrôlés plutôt que supposés : la ligne
    d'en-tête est celle qui commence par « Code commune INSEE », la colonne est
    celle dont l'en-tête est exactement « APL aux médecins généralistes », et
    la ligne d'unité doit dire « par habitant standardisé ».
    """
    import openpyxl

    livre = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True)
    lignes: list[tuple[str, str, float]] = []
    ecartees: dict[str, str] = {}
    feuilles = [n for n in livre.sheetnames if re.fullmatch(r"APL \d{4}", n)]
    if not feuilles:
        raise ValueError("aucune feuille « APL <année> » : le classeur a changé de forme")
    for nom in feuilles:
        millesime = nom.split()[1]
        feuille = livre[nom]
        rangees = feuille.iter_rows(values_only=True)
        colonne = None
        for rangee in rangees:
            if rangee and str(rangee[0] or "").strip() == "Code commune INSEE":
                cibles = [
                    i for i, cellule in enumerate(rangee)
                    if str(cellule or "").strip() == "APL aux médecins généralistes"
                ]
                if not cibles:
                    raise ValueError(f"{nom} : colonne des généralistes introuvable")
                colonne = cibles[0]
                pops = [
                    i for i, cellule in enumerate(rangee)
                    if str(cellule or "").strip().startswith("Population standardisée")
                ]
                if not pops:
                    raise ValueError(
                        f"{nom} : colonne de population standardisée introuvable —"
                        " sans elle, Paris, Lyon et Marseille (publiées par"
                        " arrondissement) ne peuvent pas s'agréger"
                    )
                colonne_pop = pops[0]
                unites = next(rangees, ())
                unite = str(unites[colonne] or "") if len(unites) > colonne else ""
                if "par habitant standardisé" not in unite:
                    raise ValueError(
                        f"{nom} : l'unité annoncée a changé ({unite[:60]!r}) —"
                        " le chiffre ne serait plus des consultations par habitant"
                    )
                break
        if colonne is None:
            raise ValueError(f"{nom} : en-tête « Code commune INSEE » introuvable")
        for rangee in rangees:
            code = str(rangee[0] or "").strip()
            if len(code) != 5:
                continue
            brut = rangee[colonne]
            if brut is None or str(brut).strip() in ("", "NS", "nd", "ND"):
                continue
            try:
                valeur = float(str(brut).replace(",", "."))
            except ValueError:
                ecartees[f"{code}/{millesime}"] = f"valeur illisible : {str(brut)[:20]}"
                continue
            if valeur < 0 or valeur > PLAFOND_PLAUSIBLE:
                ecartees[f"{code}/{millesime}"] = f"hors plage plausible : {valeur}"
                continue
            try:
                pop = float(str(rangee[colonne_pop]).replace(",", "."))
            except (TypeError, ValueError):
                pop = None
            lignes.append((code, millesime, valeur, pop))
    return lignes, ecartees


def agreger_arrondissements(
    lignes: list[tuple[str, str, float, float | None]],
) -> list[tuple[str, str, float]]:
    """Paris, Lyon et Marseille n'existent que par arrondissement dans le
    classeur : la commune s'obtient par moyenne pondérée par la population
    standardisée — la méthode d'agrégation que la DREES documente elle-même
    en tête de feuille. Un arrondissement sans population ne peut pas être
    pondéré : la commune n'est agrégée que si tous les siens en ont une."""
    sorties: list[tuple[str, str, float]] = []
    meres: dict[tuple[str, str], list[tuple[float, float | None]]] = {}
    for code, millesime, valeur, pop in lignes:
        mere = commune_mere(code)
        if mere is None:
            sorties.append((code, millesime, valeur))
        else:
            meres.setdefault((mere, millesime), []).append((valeur, pop))
    for (mere, millesime), morceaux in meres.items():
        if any(pop is None or pop <= 0 for _, pop in morceaux):
            continue
        total = sum(pop for _, pop in morceaux)
        sorties.append(
            (mere, millesime, round(sum(v * pop for v, pop in morceaux) / total, 3))
        )
    return sorties


def run(store_spec: str) -> int:
    conn = entrepot.connect()
    store = make_store(store_spec)
    declarer(conn)
    run_id = entrepot.start_run(conn, DATASET, "manual")
    try:
        avant = garde_fou_volume(conn)
        url = url_courante()
        contenu = fetch(url, timeout=600).content
        entrepot.record_asset(
            conn, store, run_id, DATASET, SOURCE, "apl-generalistes.xlsx", contenu, url,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        lignes, ecartees = lire(contenu)
        if not lignes:
            raise ValueError("aucune valeur d'APL lue : le classeur a dû changer")
        candidates = [
            (INDICATEUR, "commune", code, millesime, valeur)
            for code, millesime, valeur in agreger_arrondissements(lignes)
        ]
        gardees, hors_referentiel = filtrer_territoires_connus(conn, candidates)
        ecrites, revisees = revisions.remplacer(
            conn, run_id, [INDICATEUR], ("value",),
            ((cle, niveau, code, MILLESIME, periode, valeur)
             for cle, niveau, code, periode, valeur in gardees),
        )
        conn.execute(
            """
            insert into meta.data_quality_checks
                (run_id, dataset_id, check_name, severity, passed, observed)
            values (?, ?, 'valeurs_dans_la_plage_plausible', ?, ?, ?)
            """,
            (run_id, DATASET, "warning" if ecartees else "info", not ecartees,
             json.dumps({"ecartees": dict(list(ecartees.items())[:20]), "total": len(ecartees)})),
        )
        conn.commit()
        entrepot.finish_run(conn, run_id, "success", rows_read=len(lignes), rows_written=ecrites)
        apres = entrepot.taille(conn)
        millesimes = sorted({m for _, m, _, _ in lignes})
        print(
            f"santé : {ecrites} observations d'APL sur {millesimes[0]}-{millesimes[-1]},"
            f" {len(ecartees)} écartées, {hors_referentiel} hors référentiel,"
            f" {revisees} valeurs révisées"
        )
        print(f"volume base : {avant // 1024 // 1024} -> {apres // 1024 // 1024} Mo")
        return 0
    except Exception as error:  # noqa: BLE001 — tout échec finit tracé dans le lineage
        entrepot.finish_run(conn, run_id, "failed", error=str(error))
        raise
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default=".snapshots")
    return run(parser.parse_args().store)


if __name__ == "__main__":
    raise SystemExit(main())
