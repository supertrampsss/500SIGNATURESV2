"""Population carcérale et densité par établissement (ministère de la Justice, DGAP).

Combien de personnes sont détenues dans les prisons d'ici, et pour combien de
places : la photo mensuelle de la statistique des établissements pénitentiaires
(DGAP, sous-direction de la statistique et des études), établissement par
établissement, rattachée à la **commune d'implantation** de la prison.

**Tous les mois, dans le même run.** `revisions.remplacer` purge l'indicateur
entier avant de réécrire : charger un seul mois effacerait les autres. Le run
charge donc l'historique complet au format xlsx — chaque mois depuis le
1er janvier 2025, première parution sous ce format ; avant, la statistique
n'existe qu'en PDF, hors périmètre. Un contrôle de série borne le mouvement du
total France d'un mois au suivant : une rupture de lecture — un classeur de
2025 lu avec les yeux de 2026 — se voit là, pas dans les contrôles internes à
chaque classeur.

**La commune d'implantation, pas la commune d'origine.** Le chiffre d'une
commune décrit la prison qui s'y trouve, jamais d'où viennent les détenus.
Gradignan porte les détenus du CP de Bordeaux-Gradignan ; Bordeaux n'en porte
aucun. Le fichier ne donne pas la commune : le rattachement vient d'un
crosswalk versionné (`infra/seed/etablissements_penitentiaires.csv`), construit
une fois depuis l'annuaire de l'administration (DILA) et vérifié à la main.
Un établissement absent du crosswalk **bloque le run** : une prison nouvelle est
une décision de maintenance, pas une ligne perdue en silence.

**La densité est un ratio de sommes, jamais une moyenne de densités.** À chaque
niveau, elle rapporte la somme des détenus à la somme des places
opérationnelles. Moyenner les taux d'établissements donnerait le même poids à
une maison d'arrêt de 50 places et à Fleury-Mérogis.

**L'URL du fichier n'est pas constructible.** Le dossier de l'URL est le mois de
*mise en ligne*, pas celui des données (le fichier de février 2026 vit dans
`/2026-03/`). Le téléchargement officiel est la page annuelle de la rubrique
« statistiques mensuelles » — une page par année, chacune avec son slug — :
chaque page est lue, archivée, et les liens xlsx y sont retrouvés par leur nom
(`…_01MMAAAA.xlsx`, où MMAAAA est le mois des données). Le mois annoncé par le
nom du fichier est recoupé avec la date écrite dans le classeur. Une seule
parution déroge au nom canonique — mai 2025, publié
`Statistique_des_etablissements_2025-05-01.xlsx` — et a son motif dédié.

**Ce que les totaux publiés portent et que les lignes n'ont pas.** Quelques
quartiers sont publiés « NC » (effectif non communiqué) et quelques détenus
n'apparaissent dans aucune ligne d'établissement (34 à la DISP Rennes au
1er juillet 2026) : les totaux par direction interrégionale (Tab8) les
comptent. La somme des lignes lisibles retombe donc *en dessous* des totaux
publiés — 89 283 détenus contre 89 446 France entière au 1er juillet 2026,
0,18 % d'écart. Le contrôle par DISP exige l'égalité exacte des places et borne
ce manque à detenus ; l'écart est publié dans les contrôles qualité, jamais
réparti.

Usage : python -m plateforme.normalize.prisons [--store r2:plateforme-raw]
"""

import argparse
import csv
import io
import json
import re
from collections.abc import Sequence
from pathlib import Path

from plateforme import couverture, entrepot, revisions
from plateforme.http import telecharger
from plateforme.limites import garde_fou_volume
from plateforme.normalize.geo import MILLESIME, make_store
from plateforme.normalize.ofgl import filtrer_territoires_connus

DATASET = "justice-detenus-etablissements"
SOURCE = "justice-sdse"
THEME = "justice"

# Les pages annuelles de la rubrique, une par année de données : « -9 » porte
# 2025 (première année au format xlsx), « -11 » porte 2026. Quand le ministère
# ouvrira la page 2027, c'est ici qu'il faudra ajouter le nouveau slug — les
# liens xlsx de chaque page sont retrouvés dedans. Le slug « -10 », lui, porte
# les archives 2022 en PDF seulement : hors périmètre.
PAGES = (
    "https://www.justice.gouv.fr/documentation/etudes-et-statistiques/"
    "statistiques-mensuelles-population-detenue-ecrouee-9",
    "https://www.justice.gouv.fr/documentation/etudes-et-statistiques/"
    "statistiques-mensuelles-population-detenue-ecrouee-11",
)

# Le nom du fichier porte le mois des données (01MMAAAA : situation au 1er du
# mois) ; le dossier de l'URL porte le mois de mise en ligne, qui peut différer.
# C'est le nom qui départage les parutions laissées côte à côte sur la page.
LIEN = re.compile(
    r'href="(https://www\.justice\.gouv\.fr/sites/default/files/[^"]+/'
    r"statistique_etablissements_personnes_ecrouees_01(\d{2})(\d{4})\.xlsx)\""
)

# La seule parution qui déroge au nom canonique : mai 2025 est publié
# « Statistique_des_etablissements_2025-05-01.xlsx » (constaté le 8 août 2026
# sur la page 2025 ; l'URL canonique répond 404). Le lien reste retrouvé dans
# la page, jamais construit, et le mois du nom reste recoupé avec la date
# écrite dans le classeur.
LIEN_ALIAS = re.compile(
    r'href="(https://www\.justice\.gouv\.fr/sites/default/files/[^"]+/'
    r"Statistique_des_etablissements_(\d{4})-(\d{2})-01\.xlsx)\""
)

# Le crosswalk figé établissement -> commune d'implantation. Construit une fois
# (annuaire DILA + résolutions manuelles), jamais recalculé ici.
CROSSWALK = Path(__file__).resolve().parents[3] / "infra/seed/etablissements_penitentiaires.csv"

# Un onglet par direction interrégionale, et le total publié qui lui répond
# dans Tab8. Le dictionnaire est écrit plutôt que déduit de l'ordre des
# onglets : une DISP renommée ou déplacée doit bloquer, pas glisser.
ONGLETS_DISP = {
    "Tab14": "DISP BORDEAUX",
    "Tab15": "DISP DIJON",
    "Tab16": "DISP LILLE",
    "Tab17": "DISP LYON",
    "Tab18": "DISP MARSEILLE",
    "Tab19": "DISP PARIS",
    "Tab20": "DISP RENNES",
    "Tab21": "DISP STRASBOURG",
    "Tab22": "DISP TOULOUSE",
    "Tab23": "DSPOM",
}
ONGLET_TOTAUX = "Tab8"
FRANCE = "Total France entière"

# L'en-tête exact des onglets par établissement. Une colonne renommée ou
# déplacée arrête la lecture au lieu de servir des chiffres décalés.
EN_TETE = (
    "Etablissement",
    "Quartier (1)",
    "Capacité norme circulaire",
    "Capacité opérationnelle",
    "Ecroués détenus",
    "Densité carcérale",
)

# Ce que la colonne « Ecroués détenus » a le droit de porter d'autre qu'un
# nombre : NC, effectif non communiqué — ou une cellule vide, qui dit la même
# chose sans le dire (CP Château-Thierry et CP Maubeuge au 1er juin 2025,
# CP Sud Francilien et MA Versailles au 1er décembre 2025 : effectif et
# densité vides, quartier compté dans le total publié de la DISP). La densité
# publie en plus « -- » (capacité nulle, personne) et « Inf » (capacité nulle,
# des détenus quand même — CP Vendin-le-Vieil au 1er juillet 2026).
NON_COMMUNIQUE = "NC"
DENSITE_SANS_OBJET = {"--", "Inf", NON_COMMUNIQUE, ""}

# La date écrite dans chaque onglet, recoupée avec le mois du nom de fichier.
DATE_CLASSEUR = re.compile(r"Effectifs actualisés au 1er (\w+) (\d{4})")
MOIS_FRANCAIS = {
    "janvier": "01", "février": "02", "mars": "03", "avril": "04", "mai": "05",
    "juin": "06", "juillet": "07", "août": "08", "septembre": "09",
    "octobre": "10", "novembre": "11", "décembre": "12",
}

# La densité publiée est arrondie au dixième de point : la recalculer depuis
# détenus / places doit retomber à ±0,1 point, sans quoi une colonne a glissé.
TOLERANCE_DENSITE = 0.1

# Part des détenus d'une DISP que ses lignes d'établissement peuvent ne pas
# porter (quartiers NC + détenus non ventilés). Mesuré sur les 19 classeurs de
# janvier 2025 à juillet 2026 : 1,3 à 1,4 % à Toulouse chaque mois (quartiers
# NC), et un pic à 2,46 % à Lille au 1er juillet 2025 — le quartier CD de
# Maubeuge (198 places) passe NC en juin et juillet. Au-delà de 3 %, ce n'est
# plus du secret ponctuel, c'est une lecture amputée.
TOLERANCE_DISP = 0.03

# La somme France entière des lignes lisibles contre le total publié : 0,18 %
# d'écart mesuré (NC + non ventilés), tolérance 0,5 %.
TOLERANCE_NATIONALE = 0.005

# Témoin figé, relevé à la main sur le fichier du 1er juillet 2026 : le total
# France entière publié en Tab8. Il n'est exigé que si c'est ce fichier-là qui
# est chargé ; un fichier plus récent est contrôlé par les bornes d'ordre de
# grandeur ci-dessous.
TEMOIN_PERIODE = "2026-07"
TEMOIN_DETENUS = 89_446
TEMOIN_PLACES = 63_289

# Ordre de grandeur du parc carcéral français. Un total publié hors de ces
# bornes ne peut pas être la France : c'est une ligne lue à la place d'une
# autre.
BORNES_DETENUS = (80_000, 100_000)
BORNES_PLACES = (60_000, 70_000)

# D'un mois au suivant, le total France publié ne saute pas : la plus forte
# variation mesurée de janvier 2025 à juillet 2026 est +1,17 % (mars à avril
# 2026). Un classeur mal lu — une ligne prise pour une autre — casse cette
# continuité bien avant de sortir des bornes d'ordre de grandeur.
TOLERANCE_SERIE = 0.03

DETENUS = "justice_personnes_detenues"
DENSITE = "justice_densite_carcerale"

INDICATEURS = {
    DETENUS: {
        "libelle": "Personnes détenues",
        "unite": "count",
        "additive": True,
        "confiance": "observed",
        "public": "Les personnes écrouées et détenues dans les prisons du"
        " territoire, au 1er du mois. Le chiffre est porté par la commune où"
        " la prison est implantée, pas par celle d'où viennent les détenus."
        " N'inclut pas les écroués non détenus, sous bracelet électronique"
        " notamment.",
        "formule": "Somme des écroués détenus des quartiers des établissements"
        " implantés sur le territoire",
        "notes": "personnes écrouées détenues au 1er du mois",
    },
    DENSITE: {
        "libelle": "Densité carcérale",
        "unite": "percent",
        "additive": False,
        "confiance": "computed",
        "public": "Le nombre de détenus pour cent places opérationnelles dans"
        " les prisons du territoire, au 1er du mois. Au-dessus de 100 %,"
        " les détenus sont plus nombreux que les places. Le taux rapporte les"
        " sommes du territoire, il ne moyenne pas des taux d'établissements.",
        "formule": "Écroués détenus / places opérationnelles × 100, chaque"
        " terme sommé au niveau affiché",
        "notes": "détenus pour 100 places opérationnelles",
    },
}

TECHNIQUE = (
    "Statistique mensuelle des établissements pénitentiaires — personnes"
    " écrouées (ministère de la Justice, DGAP, sous-direction de la statistique"
    " et des études, source GENESIS), onglets par direction interrégionale"
    " (Tab14 à Tab23) du classeur xlsx, situation au 1er du mois. Le champ est"
    " celui des **écroués détenus** : les écroués non détenus — détention à"
    " domicile sous surveillance électronique, placement extérieur non"
    " hébergé — n'y sont pas (108 636 écroués dont 89 446 détenus au"
    " 1er juillet 2026). La densité rapporte les détenus à la **capacité"
    " opérationnelle** (places effectivement disponibles), et chaque niveau la"
    " calcule sur ses propres sommes — ce n'est ni une moyenne des densités"
    " d'établissements, ni forcément la densité nationale publiée par ailleurs,"
    " qui selon les publications rapporte d'autres champs (écroués détenus en"
    " établissement, hors semi-liberté…) à d'autres capacités. Chaque"
    " établissement est rattaché à sa **commune d'implantation** par un"
    " crosswalk versionné (infra/seed/etablissements_penitentiaires.csv),"
    " construit depuis la Base de données locales de l'annuaire de"
    " l'administration (DILA, licence ouverte) et vérifié à la main : le"
    " chiffre d'une commune décrit la prison qui s'y trouve, pas la population"
    " qui en vient. Quelques quartiers sont publiés « NC » (effectif non"
    " communiqué) et quelques détenus ne sont ventilés dans aucune ligne"
    " d'établissement : ils comptent dans les totaux publiés par la source"
    " mais pas ici, et ne sont jamais répartis — 163 détenus (0,18 %) au"
    " 1er juillet 2026, l'écart est publié dans les contrôles qualité. Le"
    " niveau pays somme la France entière, établissements des collectivités"
    " d'outre-mer compris (Polynésie française, Nouvelle-Calédonie, Wallis,"
    " Saint-Pierre-et-Miquelon), qui n'apparaissent à aucune maille communale"
    " ou départementale faute de territoire au référentiel. La série mensuelle"
    " court du 1er janvier 2025 — première parution au format xlsx — au dernier"
    " mois publié, sans trou, et le run la recharge entière ; les mois"
    " antérieurs à 2025 n'existent qu'en PDF et ne sont pas chargés. Licence :"
    " le fichier ne porte pas de licence propre ; les mentions légales de"
    " justice.gouv.fr placent les contenus du site sous licence ouverte 2.0"
    " (Etalab) par défaut. L'URL du fichier n'est pas constructible (le dossier"
    " est le mois de mise en ligne, pas celui des données) : les pages"
    " annuelles de la rubrique sont lues et archivées, et chaque lien xlsx y"
    " est retrouvé par son nom, qui porte le mois des données."
)


class DonneesCarceralesIncoherentes(RuntimeError):
    """Les chiffres lus ne se recoupent plus — entre eux, ou avec le témoin."""


class EtablissementInconnu(RuntimeError):
    """Un établissement du fichier n'est pas au crosswalk : maintenance requise."""


def trouver_urls(pages: Sequence[str]) -> list[tuple[str, str]]:
    """-> [(URL xlsx, période « AAAA-MM »)], tous les mois, du plus ancien au récent.

    Les liens sont retrouvés dans les pages plutôt que construits : le dossier
    de l'URL change avec le mois de mise en ligne, seul le nom du fichier dit
    le mois des données. Deux fichiers différents pour le même mois seraient
    une page en travaux : lecture refusée plutôt qu'un choix silencieux.
    """
    par_periode: dict[str, str] = {}
    for page in pages:
        trouves = [(adresse, f"{annee}-{mois}") for adresse, mois, annee in LIEN.findall(page)]
        trouves += [
            (adresse, f"{annee}-{mois}") for adresse, annee, mois in LIEN_ALIAS.findall(page)
        ]
        for adresse, periode in trouves:
            if par_periode.get(periode, adresse) != adresse:
                raise ValueError(
                    f"deux fichiers différents pour {periode} :"
                    f" {par_periode[periode]} et {adresse} — la page est en"
                    " travaux, il faut regarder avant de recharger"
                )
            par_periode[periode] = adresse
    if not par_periode:
        raise ValueError(
            f"aucun fichier « statistique_etablissements_personnes_ecrouees_01MMAAAA.xlsx »"
            f" sur {PAGES} : le ministère a renommé son fichier ou déplacé les"
            " pages (le slug change chaque année), il faut regarder avant de recharger"
        )
    return [(par_periode[periode], periode) for periode in sorted(par_periode)]


def charger_crosswalk(chemin: Path = CROSSWALK) -> dict[str, str]:
    """-> {libellé établissement du xlsx: code commune INSEE}, figé et vérifié."""
    lignes = [
        ligne for ligne in chemin.read_text(encoding="utf-8").splitlines()
        if ligne and not ligne.startswith("#")
    ]
    lecteur = csv.DictReader(io.StringIO("\n".join(lignes)))
    crosswalk = {}
    for rang in lecteur:
        etablissement, code = rang["etablissement"], rang["code_commune"]
        if etablissement in crosswalk:
            raise ValueError(f"établissement en double au crosswalk : {etablissement}")
        if not re.fullmatch(r"\d{5}|2[AB]\d{3}", code):
            raise ValueError(f"code commune invalide au crosswalk : {etablissement} -> {code}")
        crosswalk[etablissement] = code
    return crosswalk


def _nombre(valeur) -> float:
    """« 2 455 » (espace, insécable ou fine) ou 5456.0 -> float. Lève sinon."""
    texte = str(valeur).replace(" ", "").replace("\xa0", "").replace(" ", "")
    return float(texte)


def _densite(valeur) -> float | None:
    """« 205,6 % » -> 205.6 ; NC, --, Inf et vide -> None. Lève sur le reste."""
    texte = "" if valeur is None else str(valeur).strip()
    if texte in DENSITE_SANS_OBJET:
        return None
    if not texte.endswith("%"):
        raise ValueError(f"densité illisible : {valeur!r}")
    return float(texte.rstrip("%").replace(",", ".").replace("\xa0", "").replace(" ", ""))


def _periode_onglet(rangees: list[tuple]) -> str:
    """La date « Effectifs actualisés au 1er juillet 2026 » -> « 2026-07 »."""
    for rangee in rangees[:6]:
        trouve = DATE_CLASSEUR.search(str(rangee[0] or ""))
        if trouve:
            mois = MOIS_FRANCAIS.get(trouve.group(1).lower())
            if mois:
                return f"{trouve.group(2)}-{mois}"
    raise ValueError(
        "aucune ligne « Effectifs actualisés au 1er … » en tête d'onglet :"
        " impossible de dater les chiffres, le classeur a changé de forme"
    )


def lire(contenu: bytes) -> dict:
    """Classeur DGAP -> {periode, quartiers, totaux, france}.

    `quartiers` : une ligne par établissement × quartier —
    (onglet, établissement, quartier, places opérationnelles, détenus ou None
    si NC, densité publiée ou None). Les lignes « Total » par quartier et le
    total de la DISP ne sont pas des données : le premier est recalculable, le
    second sert de contrôle via Tab8. `totaux` : Tab8, par DISP. `france` :
    la ligne Total France entière de Tab8.
    """
    import openpyxl

    livre = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
    manquants = [o for o in (*ONGLETS_DISP, ONGLET_TOTAUX) if o not in livre.sheetnames]
    if manquants:
        raise ValueError(
            f"onglets {manquants} absents du classeur {livre.sheetnames[:8]}… :"
            " la publication a changé de forme, il faut regarder avant de recharger"
        )

    periodes = set()
    quartiers: list[tuple] = []
    for onglet in ONGLETS_DISP:
        rangees = list(livre[onglet].iter_rows(values_only=True))
        periodes.add(_periode_onglet(rangees))
        entetes = [str(c).strip() if c is not None else "" for c in rangees[6][:6]]
        if tuple(entetes) != EN_TETE:
            raise ValueError(
                f"{onglet} : en-tête {entetes} au lieu de {list(EN_TETE)} —"
                " les colonnes ont bougé, lecture refusée"
            )
        for rangee in rangees[7:]:
            etablissement = rangee[0]
            if etablissement is None or etablissement in ("Total", "Direction Interrégionale"):
                continue
            if str(etablissement).startswith(("(1)", "Capacité", "Densité")):
                break  # les renvois de bas d'onglet
            # Une cellule vide dit NC sans le dire (juin et décembre 2025).
            brut = "" if rangee[4] is None else str(rangee[4]).strip()
            detenus = None if brut in ("", NON_COMMUNIQUE) else _nombre(rangee[4])
            quartiers.append((
                onglet, str(etablissement), str(rangee[1]),
                _nombre(rangee[3]), detenus, _densite(rangee[5]),
            ))

    totaux, france = {}, None
    for rangee in livre[ONGLET_TOTAUX].iter_rows(values_only=True):
        libelle = str(rangee[0] or "").strip()
        if libelle in ONGLETS_DISP.values():
            totaux[libelle] = (_nombre(rangee[2]), _nombre(rangee[3]))
        elif libelle == FRANCE:
            france = (_nombre(rangee[2]), _nombre(rangee[3]))
    if france is None or set(totaux) != set(ONGLETS_DISP.values()):
        raise ValueError(
            f"{ONGLET_TOTAUX} : directions {sorted(totaux)} et France"
            f" {'présente' if france else 'absente'} — le tableau de contrôle a changé"
        )
    if len(periodes) != 1:
        raise ValueError(
            f"les onglets ne portent pas tous la même date : {sorted(periodes)}"
        )
    return {
        "periode": periodes.pop(),
        "quartiers": quartiers,
        "totaux": totaux,
        "france": {"places": france[0], "detenus": france[1]},
    }


def controler_densites(quartiers: list[tuple]) -> int:
    """Densité publiée = détenus / places recalculé, ±0,1 pt par ligne.

    C'est le contrôle qui voit une colonne décalée : capacité et détenus
    intervertis laissent toutes les sommes cohérentes entre elles, jamais ce
    ratio. -> nombre de lignes vérifiées.
    """
    verifiees = 0
    for onglet, etablissement, quartier, places, detenus, publiee in quartiers:
        if detenus is None or publiee is None or not places:
            continue
        recalculee = 100 * detenus / places
        if abs(recalculee - publiee) > TOLERANCE_DENSITE:
            raise DonneesCarceralesIncoherentes(
                f"{etablissement} ({quartier}, {onglet}) : densité publiée"
                f" {publiee:.1f} %, recalculée {recalculee:.1f} % depuis"
                f" {detenus:.0f} détenus / {places:.0f} places. Une colonne a glissé."
            )
        verifiees += 1
    if not verifiees:
        raise DonneesCarceralesIncoherentes("aucune densité vérifiable : rien n'a été lu")
    return verifiees


def controler_disp(quartiers: list[tuple], totaux: dict) -> dict[str, float]:
    """Chaque DISP retombe sur Tab8 : places exactes, détenus bornés.

    Les places opérationnelles des lignes d'établissement — quartiers NC
    compris — somment exactement au total publié. Les détenus des lignes
    lisibles peuvent manquer le total (quartiers NC, détenus non ventilés) :
    le manque doit rester positif et sous TOLERANCE_DISP — au-delà, c'est une
    lecture amputée, que rien d'interne ne verrait autrement.
    -> détenus manquants par DISP, à publier dans les contrôles qualité.
    """
    manquants = {}
    for onglet, disp in ONGLETS_DISP.items():
        lignes = [q for q in quartiers if q[0] == onglet]
        if not lignes:
            raise DonneesCarceralesIncoherentes(f"{disp} : aucune ligne d'établissement lue")
        places, places_publiees = sum(q[3] for q in lignes), totaux[disp][0]
        if places != places_publiees:
            raise DonneesCarceralesIncoherentes(
                f"{disp} : {places:.0f} places opérationnelles en somme des"
                f" établissements contre {places_publiees:.0f} publiées en"
                f" {ONGLET_TOTAUX}. Un établissement a été perdu ou compté deux fois."
            )
        detenus = sum(q[4] for q in lignes if q[4] is not None)
        manque = totaux[disp][1] - detenus
        if manque < 0 or manque > TOLERANCE_DISP * totaux[disp][1]:
            raise DonneesCarceralesIncoherentes(
                f"{disp} : {detenus:.0f} détenus en somme des établissements contre"
                f" {totaux[disp][1]:.0f} publiés — {manque:+.0f}, hors de ce que les"
                f" quartiers NC et les non-ventilés expliquent"
                f" (0 à {100 * TOLERANCE_DISP:.0f} %)."
            )
        manquants[disp] = manque
    return manquants


def controler_national(quartiers: list[tuple], france: dict, periode: str) -> dict:
    """Le total France entière : bornes, témoin figé, et somme des lignes.

    Le témoin du 1er juillet 2026 est exigé à l'identique si c'est ce
    fichier-là qui est chargé ; un fichier plus récent passe par les bornes
    d'ordre de grandeur. Dans les deux cas, la somme des lignes lisibles doit
    retrouver le total publié à TOLERANCE_NATIONALE près — c'est le seul
    contrôle qu'une lecture tronquée d'onglets entiers ne peut pas passer.
    """
    detenus_publies, places_publiees = france["detenus"], france["places"]
    if periode == TEMOIN_PERIODE:
        if (detenus_publies, places_publiees) != (TEMOIN_DETENUS, TEMOIN_PLACES):
            raise DonneesCarceralesIncoherentes(
                f"fichier du {TEMOIN_PERIODE} : {detenus_publies:.0f} détenus /"
                f" {places_publiees:.0f} places publiés France entière, témoin"
                f" relevé à la main {TEMOIN_DETENUS} / {TEMOIN_PLACES}."
            )
    elif not (
        BORNES_DETENUS[0] <= detenus_publies <= BORNES_DETENUS[1]
        and BORNES_PLACES[0] <= places_publiees <= BORNES_PLACES[1]
    ):
        raise DonneesCarceralesIncoherentes(
            f"France entière : {detenus_publies:.0f} détenus /"
            f" {places_publiees:.0f} places publiés, hors des bornes"
            f" {BORNES_DETENUS} / {BORNES_PLACES} — ce n'est pas le parc carcéral"
            " français, une ligne a été lue à la place d'une autre."
        )
    detenus = sum(q[4] for q in quartiers if q[4] is not None)
    ecart = abs(detenus - detenus_publies) / detenus_publies
    if ecart > TOLERANCE_NATIONALE:
        raise DonneesCarceralesIncoherentes(
            f"somme des lignes lisibles : {detenus:.0f} détenus contre"
            f" {detenus_publies:.0f} publiés France entière, {100 * ecart:.2f} %"
            f" d'écart pour {100 * TOLERANCE_NATIONALE:.1f} % tolérés."
            " Les quartiers NC n'expliquent pas ça : lecture tronquée."
        )
    return {
        "detenus_publies": detenus_publies,
        "places_publiees": places_publiees,
        "detenus_lus": detenus,
        "ecart_relatif": round(ecart, 5),
    }


def _mois_suivant(periode: str) -> str:
    """« 2025-12 » -> « 2026-01 »."""
    annee, mois = int(periode[:4]), int(periode[5:])
    return f"{annee + mois // 12}-{mois % 12 + 1:02d}"


def controler_serie(serie: dict[str, float]) -> dict:
    """La série des totaux France se tient : aucun trou, aucun saut.

    Chaque contrôle par classeur compare un classeur à lui-même : un fichier de
    2025 lu de travers mais cohérent avec lui-même les passerait tous. D'un
    mois au suivant, le parc carcéral bouge de moins d'un point et demi
    (mesuré 2025-2026) : un total qui saute de plus de TOLERANCE_SERIE est une
    rupture de lecture, pas une vague d'incarcérations. Et un mois manquant
    n'est pas un cas à absorber : le remplacement couvrant l'indicateur entier,
    un mois disparu des pages annuelles disparaîtrait de l'entrepôt en silence.
    -> {mois contrôlés, variation maximale, où} pour les contrôles qualité.
    """
    periodes = sorted(serie)
    variation_max, entre = 0.0, None
    for avant, apres in zip(periodes, periodes[1:], strict=False):
        if apres != _mois_suivant(avant):
            raise DonneesCarceralesIncoherentes(
                f"trou dans la série mensuelle : rien entre {avant} et {apres}."
                " Un mois publié a quitté les pages annuelles ou son lien a"
                " changé de nom — recharger sans lui l'effacerait de l'entrepôt."
            )
        variation = abs(serie[apres] - serie[avant]) / serie[avant]
        if variation > TOLERANCE_SERIE:
            raise DonneesCarceralesIncoherentes(
                f"de {avant} à {apres}, le total France passe de"
                f" {serie[avant]:.0f} à {serie[apres]:.0f} détenus"
                f" ({100 * variation:.1f} % pour {100 * TOLERANCE_SERIE:.0f} %"
                " tolérés) : un des deux classeurs a été lu de travers."
            )
        if variation > variation_max:
            variation_max, entre = variation, f"{avant} -> {apres}"
    return {
        "mois": len(periodes),
        "variation_maximale": round(variation_max, 4),
        "entre": entre,
    }


def controler_crosswalk(quartiers: list[tuple], crosswalk: dict[str, str]) -> int:
    """Chaque établissement du fichier trouve sa commune, ou rien ne sort.

    Un établissement nouveau — ouverture, renommage — doit être ajouté au
    crosswalk versionné après vérification de sa commune d'implantation :
    c'est une maintenance consciente, pas un cas à absorber.
    """
    etablissements = {q[1] for q in quartiers}
    inconnus = sorted(etablissements - set(crosswalk))
    if inconnus:
        raise EtablissementInconnu(
            f"{len(inconnus)} établissement(s) absent(s) de"
            f" {CROSSWALK.name} : {inconnus}. Prison nouvelle ou renommée —"
            " vérifier sa commune d'implantation et compléter le crosswalk."
        )
    return len(etablissements)


def _departement(code: str) -> str | None:
    """Commune -> département, ou None pour les collectivités d'outre-mer.

    Les communes des COM (Saint-Pierre-et-Miquelon 975, Wallis 986, Polynésie
    987, Nouvelle-Calédonie 988) n'ont pas de département : leurs
    établissements ne comptent qu'au niveau pays.
    """
    if code[:2] in ("2A", "2B"):
        return code[:2]
    if code.startswith(("97", "98")):
        return code[:3] if code[:3] in ("971", "972", "973", "974", "976") else None
    return code[:2]


def par_commune(quartiers: list[tuple], crosswalk: dict[str, str]) -> dict[str, list[float]]:
    """-> {code commune: [détenus, places]}, quartiers NC exclus des deux.

    Un quartier dont l'effectif n'est pas communiqué sort avec ses places :
    garder ses places au dénominateur ferait baisser la densité d'une commune
    précisément là où la source ne dit pas combien de détenus s'y trouvent.
    """
    communes: dict[str, list[float]] = {}
    for _, etablissement, _, places, detenus, _ in quartiers:
        if detenus is None:
            continue
        cible = communes.setdefault(crosswalk[etablissement], [0.0, 0.0])
        cible[0] += detenus
        cible[1] += places
    return communes


def aggreger(
    communes: dict[str, list[float]], regions: dict[str, str], periode: str
) -> list[tuple]:
    """-> candidates (indicateur, niveau, code, période, valeur), sommes par niveau.

    La densité de chaque niveau est recalculée sur ses sommes — jamais une
    moyenne de densités — et n'est pas écrite quand le niveau n'a aucune place
    opérationnelle (0 détenu / 0 place : MA de Belfort au 1er juillet 2026).
    Le pays somme tout, collectivités d'outre-mer comprises ; un département
    absent de `regions` (référentiel) arrête plutôt que de perdre sa région.
    """
    sommes: dict[tuple[str, str], list[float]] = {("pays", "FR"): [0.0, 0.0]}
    for code, (detenus, places) in communes.items():
        cibles = [("commune", code)]
        departement = _departement(code)
        if departement:
            region = regions.get(departement)
            if region is None:
                raise DonneesCarceralesIncoherentes(
                    f"département {departement} sans région au référentiel :"
                    " le rattachement du crosswalk ou le référentiel a dérivé"
                )
            cibles += [("departement", departement), ("region", region)]
        cibles.append(("pays", "FR"))
        for cible in cibles:
            somme = sommes.setdefault(cible, [0.0, 0.0])
            somme[0] += detenus
            somme[1] += places

    candidates = []
    for (niveau, code), (detenus, places) in sorted(sommes.items()):
        candidates.append((DETENUS, niveau, code, periode, detenus))
        if places:
            candidates.append((DENSITE, niveau, code, periode, round(100 * detenus / places, 1)))
    return candidates


def detenus_par_maille(candidates: list[tuple]) -> dict[str, float]:
    """La couverture se mesure en détenus : une prison perdue pèse ses détenus."""
    par_maille: dict[str, float] = {}
    for indicateur, niveau, _, _, valeur in candidates:
        if indicateur == DETENUS:
            par_maille[niveau] = par_maille.get(niveau, 0.0) + valeur
    return par_maille


def declarer(conn) -> None:
    with conn.cursor() as curseur:
        for identifiant, fiche in sorted(INDICATEURS.items()):
            definition = curseur.execute(
                """
                insert into core.indicator_definitions
                    (public_definition, technical_definition, formula, unit_notes,
                     confidence_level, badges)
                values (?, ?, ?, ?, ?, array['Officiel','Donnée brute'])
                returning definition_id
                """,
                (
                    fiche["public"], TECHNIQUE, fiche["formule"], fiche["notes"],
                    fiche["confiance"],
                ),
            ).fetchone()[0]
            curseur.execute(
                """
                insert into core.indicators
                    (indicator_id, dataset_id, definition_id, theme, label_fr, unit,
                     additive, geo_levels, time_granularity, published)
                values (?, ?, ?, ?, ?, ?, ?,
                        array['commune','departement','region','pays'], 'mensuelle', true)
                on conflict (indicator_id) do update set unit = excluded.unit,
                    definition_id = excluded.definition_id, label_fr = excluded.label_fr,
                    theme = excluded.theme, additive = excluded.additive, published = true
                """,
                (
                    identifiant, DATASET, definition, THEME, fiche["libelle"],
                    fiche["unite"], fiche["additive"],
                ),
            )
        curseur.execute(
            "delete from core.indicator_definitions d where not exists"
            " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
        )
    conn.commit()


COLONNES = ("value",)


def run(store_spec: str) -> int:
    conn = entrepot.connect()
    store = make_store(store_spec)
    declarer(conn)
    run_id = entrepot.start_run(conn, DATASET, "manual")
    try:
        garde_fou_volume(conn)
        pages = []
        for adresse_page in PAGES:
            page = telecharger(adresse_page, timeout=120)
            entrepot.record_asset(
                conn, store, run_id, DATASET, SOURCE,
                f"page-{adresse_page.rsplit('/', 1)[-1]}.html",
                page, adresse_page, "text/html",
            )
            pages.append(page.decode("utf-8", errors="replace"))
        liens = trouver_urls(pages)
        crosswalk = charger_crosswalk()
        regions = dict(conn.execute(
            "select geo_code, parent_code from geo.geography_reference"
            " where geo_level = 'departement' and vintage = ?", (MILLESIME,),
        ).fetchall())

        # Tous les mois dans le même run : `revisions.remplacer` purge
        # l'indicateur entier, un chargement partiel effacerait le reste.
        candidates: list[tuple] = []
        gardees: list[tuple] = []
        ecartes: set[str] = set()
        serie: dict[str, float] = {}
        par_periode: dict[str, dict] = {}
        lignes_lues = 0
        for adresse, periode_fichier in liens:
            contenu = telecharger(adresse, timeout=600)
            entrepot.record_asset(
                conn, store, run_id, DATASET, SOURCE, adresse.rsplit("/", 1)[-1],
                contenu, adresse,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            lu = lire(contenu)
            if lu["periode"] != periode_fichier:
                raise DonneesCarceralesIncoherentes(
                    f"le nom du fichier annonce {periode_fichier}, le classeur dit"
                    f" {lu['periode']} : la page a servi un fichier sous un mauvais nom"
                )
            lignes_lues += len(lu["quartiers"])
            densites = controler_densites(lu["quartiers"])
            manquants = controler_disp(lu["quartiers"], lu["totaux"])
            national = controler_national(lu["quartiers"], lu["france"], lu["periode"])
            etablissements = controler_crosswalk(lu["quartiers"], crosswalk)
            communes = par_commune(lu["quartiers"], crosswalk)
            du_mois = aggreger(communes, regions, lu["periode"])
            gardees_mois, ecartes_mois = filtrer_territoires_connus(conn, du_mois)
            candidates += du_mois
            gardees += gardees_mois
            ecartes |= ecartes_mois
            serie[lu["periode"]] = national["detenus_publies"]
            par_periode[lu["periode"]] = {
                "etablissements": etablissements,
                "densites_verifiees": densites,
                "quartiers_nc": sum(1 for q in lu["quartiers"] if q[4] is None),
                "detenus_manquants_par_disp": {
                    disp: valeur for disp, valeur in sorted(manquants.items()) if valeur
                },
                **national,
            }
            entrepot.etape(
                f"{lu['periode']} : {len(contenu) // 1024} Ko,"
                f" {len(lu['quartiers'])} lignes,"
                f" {national['detenus_publies']:.0f} détenus publiés"
            )
        continuite = controler_serie(serie)

        ecrites, revisees = revisions.remplacer(
            conn, run_id, sorted(INDICATEURS), COLONNES,
            [
                (indicateur, niveau, code, MILLESIME, periode, valeur)
                for indicateur, niveau, code, periode, valeur in gardees
            ],
        )
        parts = couverture.controler(
            detenus_par_maille(candidates), detenus_par_maille(gardees)
        )
        conn.execute(
            """
            insert into meta.data_quality_checks
                (run_id, dataset_id, check_name, severity, passed, observed)
            values (?, ?, 'densites_disp_serie_et_temoin_national', 'blocker', true, ?)
            """,
            (run_id, DATASET, json.dumps({
                "serie": continuite,
                "par_periode": {p: par_periode[p] for p in sorted(par_periode)},
                "hors_referentiel": sorted(ecartes),
                "couverture": {n: round(p, 4) for n, p in sorted(parts.items())},
            })),
        )
        conn.commit()
        entrepot.finish_run(
            conn, run_id, "success", rows_read=lignes_lues, rows_written=ecrites
        )
        periodes = sorted(serie)
        dernier = par_periode[periodes[-1]]
        print(
            f"prisons : {ecrites} observations sur {len(periodes)} mois"
            f" ({periodes[0]} -> {periodes[-1]}), variation maximale"
            f" {100 * continuite['variation_maximale']:.2f} % ;"
            f" au {periodes[-1]}, {dernier['detenus_lus']:.0f} détenus lus pour"
            f" {dernier['detenus_publies']:.0f} publiés sur"
            f" {dernier['etablissements']} établissements ;"
            f" {len(ecartes)} territoires hors référentiel, {revisees} valeurs révisées"
        )
        return 0
    except Exception as error:  # noqa: BLE001 — tout échec finit tracé dans le lineage
        entrepot.finish_run(conn, run_id, "failed", error=str(error))
        raise
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default=".snapshots")
    return run(parser.parse_args().store)


if __name__ == "__main__":
    raise SystemExit(main())
