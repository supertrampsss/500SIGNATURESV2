"""Le budget de la Sécurité sociale ligne à ligne : le PLFSS et ses annexes.

Le site règle le budget de l'État — 594,0 Md€ de crédits de paiement au PLF
2025 — et laissait croire que c'était là que se jouait la dépense publique.
C'est faux d'un facteur trois : les régimes obligatoires de base de la sécurité
sociale portent **676,9 Md€ de charges nettes** au PLFSS 2026, pour 659,5 Md€ de
produits nets et un résultat de −17,5 Md€. Ce module charge ces deux colonnes
poste par poste, plus l'ONDAM et ses six sous-objectifs.

**Trois grandeurs, trois périmètres, et rien ne s'additionne.**

- Les **charges et produits nets** viennent de l'annexe 3 du PLFSS, tableau 5.
  Périmètre : les régimes obligatoires de base **consolidés** (colonne « Régimes
  de base »), hors FSV. Comptabilité en droits constatés, pas budgétaire.
- L'**ONDAM** vient de l'annexe 5, tableau 8. Ce n'est pas une ligne de charge :
  c'est un *objectif de dépenses* voté à part, qui traverse les branches maladie,
  AT-MP et autonomie sans être la somme d'aucune d'elles. Il vaut 270,4 Md€ en
  2026 quand les seules charges de la branche maladie font 267,5 Md€ : les deux
  ne se soustraient pas, ne s'additionnent pas, et ne se comparent pas.
- Le **budget de l'État** publié par ailleurs sur ce site est une troisième
  chose encore. Entre lui et la Sécurité sociale circulent des transferts massifs
  — TVA affectée, compensations d'exonérations, cotisations prises en charge par
  l'État — que chacun compte de son côté. Additionner les deux totaux compterait
  deux fois ces montants.

**La colonne consolidée n'est pas la somme des branches, et c'est le piège
principal.** Le tableau 5 donne cinq colonnes de branche puis une colonne
« Régimes de base ». En 2026 les cinq branches font 695 943,8 M€ de charges là où
la colonne consolidée en donne 676 925,3 : l'écart, 19,0 Md€, est celui des
transferts entre branches, comptés en charge chez l'une et en produit chez
l'autre. Seule la colonne consolidée est chargée ici — une lecture par branche
publierait un total qui n'existe nulle part.

**Le titre du tableau dit « en milliards d'euros » et les cellules sont en
millions.** La ligne « CHARGES NETTES » y vaut 676 925,28. Lire le titre plutôt
que les cellules diviserait toute la Sécurité sociale par mille sans qu'aucun
contrôle interne au tableau ne bouge : ils resteraient tous cohérents entre eux.
C'est pourquoi l'ordre de grandeur est confronté à une **autre source déjà
chargée**, les prestations de protection sociale de la DREES.

**Un seul exercice, et le refus est écrit ici.** L'annexe 3 du PLFSS 2025 publie
le même tableau arrondi au dixième de milliard : les charges diverses y valent
100 M€, les produits financiers de la branche maladie 0, et la somme des postes
s'écarte de son total de 100 M€ par pur arrondi. Un simulateur bâti là-dessus
proposerait de régler des lignes à zéro et afficherait un solde qui ne retombe
pas. L'exercice 2026, lui, est publié au millième de million et toutes les
identités y sont exactes. On s'arrête donc là où la source devient utilisable.

**Le PLFSS n'est pas la LFSS.** Ces montants sont ceux du *projet* déposé le
14 octobre 2025 ; le Parlement l'a amendé et la loi promulguée le 31 décembre
2025 porte d'autres chiffres. Ils ne se comparent ni aux comptes définitifs ni
au « trou de la Sécu » commenté après coup.

**Quatre prises de contrôle, dont deux qui ne partagent aucun calcul.**

1. *Interne* : les postes font leur total, à chaque étage et des deux côtés, et
   le résultat net est la différence des deux totaux.
2. *Croisée dans le fichier* : le tableau 4 de la même annexe construit le solde
   tout autrement — solde tendanciel plus mesures, branche par branche. Ses six
   soldes après mesures doivent retomber sur les six résultats nets du tableau 5.
   Deux constructions indépendantes de la même grandeur.
3. *Croisée entre fichiers* : les six taux d'évolution de l'ONDAM publiés dans la
   synthèse de l'annexe 5 doivent redonner les montants du tableau 8 à partir de
   leur base, et les six sous-objectifs doivent faire l'ONDAM.
4. *Croisée entre sources* : les charges nettes des régimes de base sont
   confrontées aux prestations de protection sociale de la DREES déjà chargées.
   C'est le seul contrôle que ni un changement de colonne ni un changement
   d'unité dans le fichier de la DSS ne peut satisfaire.

Usage : python -m plateforme.normalize.plfss [--store r2:plateforme-raw]
"""

import argparse
import io
import json
from collections import defaultdict
from urllib.parse import quote

import openpyxl

from plateforme import couverture, entrepot
from plateforme.http import contexte_sans_confidentialite_persistante, telecharger
from plateforme.limites import garde_fou_volume
from plateforme.normalize.geo import make_store

DATASET = "plfss-annexes"
SOURCE = "securite-sociale"
BASE = "https://www.securite-sociale.fr/files/live/sites/SSFR/files/medias/PLFSS"

LOI = "PLFSS"
MESURE_EQUILIBRE = "charges_et_produits_nets"
MESURE_OBJECTIF = "objectif_de_depenses"

# Les noms de fichier des annexes ne suivent aucune règle d'un exercice à
# l'autre — « Annexe 3 PLFSS 2026 Tableaux données .xlsx » porte une espace
# avant le point, « Annexe 5 Tableaux PLFSS 2026 (1).xlsx » un numéro de
# version. Ils sont écrits, pas devinés.
ANNEXES = {
    "2026": {
        "equilibre": "Annexe 3 PLFSS 2026 Tableaux données .xlsx",
        "ondam": "Annexe 5 Tableaux PLFSS 2026 (1).xlsx",
    },
}

# Les onglets, et la colonne de chaque tableau. Les indices sont ceux
# d'openpyxl, à partir de 1 ; l'intitulé attendu en tête de colonne est vérifié
# avant toute lecture — un tableau qui gagnerait une branche décalerait la
# colonne consolidée sans rien changer à la forme du fichier.
ONGLET_EQUILIBRE = "II.tab5"
ONGLET_SOLDES = "II.tab4"
ONGLET_ONDAM = "I.5_Tableau 8"
ONGLET_TAUX = "Synthèse_Tableau 1"

COLONNE_LIBELLE_EQUILIBRE = 3
COLONNE_CONSOLIDE = 9
ENTETE_CONSOLIDE = "Régimes de base"
# Les cinq branches, dans l'ordre du tableau 5. Elles ne sont pas publiées — la
# somme des branches n'est pas le consolidé — mais elles servent au contrôle
# croisé avec le tableau 4, qui les donne dans un autre ordre.
BRANCHES_EQUILIBRE = {"Maladie": 4, "Vieillesse": 5, "Famille": 6, "AT-MP": 7, "Autonomie": 8}
# La clé publiée de chaque branche : sans accent, sans tiret, stable dans une
# URL et dans une contrainte SQL.
CLE_BRANCHE = {
    "Maladie": "maladie",
    "Vieillesse": "vieillesse",
    "Famille": "famille",
    "AT-MP": "atmp",
    "Autonomie": "autonomie",
}

COLONNE_LIBELLE_SOLDES = 1
BRANCHES_SOLDES = {"Maladie": 2, "AT-MP": 3, "Vieillesse": 4, "Famille": 5, "Autonomie": 6}
COLONNE_SOLDES_CONSOLIDE = 7
ENTETE_SOLDES_CONSOLIDE = "RB"
LIGNE_SOLDE_APRES_MESURES = "Solde après-mesures"

COLONNE_LIBELLE_ONDAM = 2
COLONNE_ONDAM_BASE = 3
COLONNE_ONDAM_TAUX = 4
COLONNE_ONDAM_MONTANT = 5
COLONNE_TAUX_LIBELLE = 2
COLONNE_TAUX_VALEUR = 3

SECTION_CHARGES = "CHARGES NETTES"
SECTION_PRODUITS = "PRODUITS NETS"
LIGNE_RESULTAT = "RESULTAT NET"

# Le tableau 5 est en millions d'euros — quoi qu'en dise son titre. Le site
# publie en euros, comme le budget de l'État et la dette.
MILLIONS = 1_000_000
# Le tableau 8 de l'annexe 5, lui, est bien en milliards.
MILLIARDS = 1_000_000_000

# Les montants du tableau 5 sont écrits au millième de million : l'écart attendu
# entre un poste et la somme de ses composantes est nul. Un centième de million,
# soit dix mille euros, laisse passer le flottement d'un tableur et rien d'autre.
TOLERANCE_EUROS = 0.01 * MILLIONS

# L'ONDAM est publié au dixième de milliard. Un montant reconstruit depuis sa
# base et son taux porte donc deux arrondis : 0,15 Md€ est l'écart honnête
# maximal, 0,16 la tolérance. Constaté sur les six sous-objectifs 2026 : 0,08 au
# pire, sur les soins de ville.
TOLERANCE_ONDAM = 0.16 * MILLIARDS

# Les charges nettes des régimes de base rapportées aux prestations de protection
# sociale versées tous régimes confondus (DREES, `drees_protection_sociale_total`).
# Le rapport constaté est 0,73 en 2026 contre 2024 : la protection sociale couvre
# des régimes complémentaires, l'État, les collectivités et les mutuelles que les
# régimes de base ne portent pas, et les régimes de base portent des charges de
# gestion et des transferts qui ne sont pas des prestations. La fourchette est
# large parce qu'elle n'a qu'un travail : arrêter un changement d'unité ou de
# colonne, qui déplace ce rapport d'un facteur mille.
RAPPORT_DREES = (0.4, 1.1)
INDICATEUR_DREES = "drees_protection_sociale_total"


class BudgetSocialIncoherent(RuntimeError):
    """Un contrôle bloquant a refusé l'exercice."""


def normaliser(texte) -> str:
    """L'intitulé d'une ligne, ramené à ce qui l'identifie.

    La source écrit « Etablisements et services pour personnes âgées » ici et
    « Dépenses relatives aux  établisements et services pour personnes âgées »
    là, avec deux espaces au milieu. Les espaces sont donc réduits et
    l'apostrophe typographique ramenée à l'apostrophe droite : ce sont les deux
    seules variations qu'une frappe corrigée puisse introduire sans que la ligne
    change de sens.
    """
    if texte is None:
        return ""
    return " ".join(str(texte).replace("’", "'").split())


# La nomenclature publiée, écrite ligne à ligne : (section, intitulé source) ->
# (code, code du parent, niveau). C'est elle qui décide ce qui est un poste, ce
# qui est son parent, et donc ce qui se somme. Une ligne que la source
# ajouterait n'y figurerait pas : elle ne serait pas écrite, et la couverture
# lue/écrite tomberait sous son seuil — voir `run`.
#
# Les deux côtés portent chacun une ligne « Transferts nets » : la clé porte donc
# la section, sans quoi les transferts versés et les transferts reçus tomberaient
# sur le même code et l'un écraserait l'autre.
POSTES = {
    (SECTION_CHARGES, "Prestations sociales nettes"): ("D-PRE", None, "agregat"),
    (SECTION_CHARGES, "Prestations légales nettes"): ("D-PRE-LEG", "D-PRE", "poste"),
    (SECTION_CHARGES, "Prestations extralégales nettes"): ("D-PRE-XLE", "D-PRE", "poste"),
    (SECTION_CHARGES, "Transferts nets"): ("D-TRF", None, "agregat"),
    (SECTION_CHARGES, "Transferts avec d'autres régimes de base"): (
        "D-TRF-RB", "D-TRF", "poste",
    ),
    (SECTION_CHARGES, "Transfert avec des fonds"): ("D-TRF-FDS", "D-TRF", "poste"),
    (SECTION_CHARGES, "Autres transferts versés"): ("D-TRF-AUT", "D-TRF", "poste"),
    (SECTION_CHARGES, "Charges de gestion courante"): ("D-GES", None, "agregat"),
    (SECTION_CHARGES, "Autres charges"): ("D-AUT", None, "agregat"),
    (SECTION_CHARGES, "Charges financières"): ("D-AUT-FIN", "D-AUT", "poste"),
    (SECTION_CHARGES, "Charges diverses"): ("D-AUT-DIV", "D-AUT", "poste"),
    (SECTION_PRODUITS, "Cotisations, contributions et recettes fiscales nettes"): (
        "R-COT", None, "agregat",
    ),
    (SECTION_PRODUITS, "Cotisations sociales"): ("R-COT-SOC", "R-COT", "poste"),
    (SECTION_PRODUITS, "Cotisations prises en charge par l'État"): (
        "R-COT-ETA", "R-COT", "poste",
    ),
    (SECTION_PRODUITS, "Contribution de l'employeur"): ("R-COT-EMP", "R-COT", "poste"),
    (SECTION_PRODUITS, "CSG"): ("R-COT-CSG", "R-COT", "poste"),
    (SECTION_PRODUITS, "Autres contributions sociales"): ("R-COT-ACS", "R-COT", "poste"),
    (SECTION_PRODUITS, "Recettes fiscales"): ("R-COT-FIS", "R-COT", "poste"),
    (SECTION_PRODUITS, "Charges liées au non-recouvrement"): ("R-COT-NRE", "R-COT", "poste"),
    (SECTION_PRODUITS, "Sur cotisations"): ("R-COT-NRE-COT", "R-COT-NRE", "sous_poste"),
    (SECTION_PRODUITS, "Sur CSG activité"): ("R-COT-NRE-CSG", "R-COT-NRE", "sous_poste"),
    (SECTION_PRODUITS, "Sur recettes fiscales et autres contributions sociales"): (
        "R-COT-NRE-FIS", "R-COT-NRE", "sous_poste",
    ),
    (SECTION_PRODUITS, "Transferts nets"): ("R-TRF", None, "agregat"),
    (SECTION_PRODUITS, "Reçus des régimes de base"): ("R-TRF-RB", "R-TRF", "poste"),
    (SECTION_PRODUITS, "Reçus des fonds de financement"): ("R-TRF-FDS", "R-TRF", "poste"),
    (SECTION_PRODUITS, "Reçus de l'Etat"): ("R-TRF-ETA", "R-TRF", "poste"),
    (SECTION_PRODUITS, "Autres"): ("R-TRF-AUT", "R-TRF", "poste"),
    (SECTION_PRODUITS, "Autres produits"): ("R-AUT", None, "agregat"),
    (SECTION_PRODUITS, "Produits financiers"): ("R-AUT-FIN", "R-AUT", "poste"),
    (SECTION_PRODUITS, "Produits divers"): ("R-AUT-DIV", "R-AUT", "poste"),
}

COTE = {SECTION_CHARGES: "depense", SECTION_PRODUITS: "recette"}

# Les intitulés publiés. Ceux de la source disent « nettes » quatre fois de
# suite sans que le lecteur sache de quoi ; ceux-ci disent la même chose en
# français, et c'est le seul endroit où la reformulation est permise.
LIBELLES = {
    "D-PRE": "Prestations sociales",
    "D-PRE-LEG": "Prestations légales",
    "D-PRE-XLE": "Prestations extralégales (action sociale)",
    "D-TRF": "Transferts versés",
    "D-TRF-RB": "Transferts vers d'autres régimes de base",
    "D-TRF-FDS": "Transferts vers des fonds",
    "D-TRF-AUT": "Autres transferts versés",
    "D-GES": "Charges de gestion courante",
    "D-AUT": "Autres charges",
    "D-AUT-FIN": "Charges financières",
    "D-AUT-DIV": "Charges diverses",
    "R-COT": "Cotisations, contributions et recettes fiscales",
    "R-COT-SOC": "Cotisations sociales",
    "R-COT-ETA": "Cotisations prises en charge par l'État",
    "R-COT-EMP": "Contribution de l'employeur (fonction publique)",
    "R-COT-CSG": "Contribution sociale généralisée (CSG)",
    "R-COT-ACS": "Autres contributions sociales",
    "R-COT-FIS": "Recettes fiscales affectées",
    "R-COT-NRE": "Non-recouvrement (se déduit)",
    "R-COT-NRE-COT": "Non-recouvrement sur cotisations",
    "R-COT-NRE-CSG": "Non-recouvrement sur CSG d'activité",
    "R-COT-NRE-FIS": "Non-recouvrement sur recettes fiscales et autres contributions",
    "R-TRF": "Transferts reçus",
    "R-TRF-RB": "Reçus d'autres régimes de base",
    "R-TRF-FDS": "Reçus de fonds de financement",
    "R-TRF-ETA": "Reçus de l'État",
    "R-TRF-AUT": "Autres transferts reçus",
    "R-AUT": "Autres produits",
    "R-AUT-FIN": "Produits financiers",
    "R-AUT-DIV": "Produits divers",
}

# Les sous-objectifs de l'ONDAM, dans l'ordre du tableau 8. Le troisième est un
# regroupement : ses deux composantes sont les sous-objectifs réellement votés,
# et ce sont elles qui font le total avec les quatre autres.
SOUS_OBJECTIFS = {
    "ONDAM TOTAL": ("O-TOTAL", None, "objectif"),
    "Soins de ville": ("O-SDV", None, "sous_objectif"),
    "Etablissements de santé": ("O-ETS", None, "sous_objectif"),
    "Etablissements et services médico-sociaux": ("O-MS", None, "sous_objectif"),
    "Dépenses relatives aux établisements et services pour personnes âgées": (
        "O-MS-PA", "O-MS", "sous_objectif",
    ),
    "Dépenses relatives aux établisements et services pour personnes handicapées": (
        "O-MS-PH", "O-MS", "sous_objectif",
    ),
    "Dépenses relatives au Fonds d'intervention régional et Soutien national à"
    " l'investissement": ("O-FIR", None, "sous_objectif"),
    "Autres prises en charges": ("O-APC", None, "sous_objectif"),
}

LIBELLES_OBJECTIF = {
    "O-TOTAL": "ONDAM",
    "O-SDV": "Soins de ville",
    "O-ETS": "Établissements de santé",
    "O-MS": "Établissements et services médico-sociaux",
    "O-MS-PA": "Établissements et services pour personnes âgées",
    "O-MS-PH": "Établissements et services pour personnes handicapées",
    "O-FIR": "Fonds d'intervention régional et investissement",
    "O-APC": "Autres prises en charge",
}

# Les six sous-objectifs qui font le total, et l'ordre dans lequel la synthèse de
# l'annexe 5 publie leurs taux d'évolution. Les intitulés y diffèrent de ceux du
# tableau 8 — « Établissement de santé » au singulier, « Etablisements » sans
# accent ni double s : le rapprochement se fait donc par position, et cette liste
# est ce qui rend cette position vérifiable.
RANG_DES_TAUX = ["O-SDV", "O-ETS", "O-MS-PA", "O-MS-PH", "O-FIR", "O-APC", "O-TOTAL"]
FEUILLES_OBJECTIF = [code for code in RANG_DES_TAUX if code != "O-TOTAL"]


def url(exercice: str, annexe: str) -> str:
    return f"{BASE}/{exercice}/{quote(ANNEXES[exercice][annexe])}"


def classeur(contenu: bytes):
    return openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)


def _nombre(valeur) -> float | None:
    return float(valeur) if isinstance(valeur, int | float) else None


def _cellule(rang: tuple, colonne: int):
    return rang[colonne - 1] if len(rang) >= colonne else None


def controler_entete(rang: tuple, colonne: int, attendu: str, onglet: str) -> None:
    """La colonne lue porte-t-elle encore le titre qu'on croit ?

    Le tableau 5 aligne cinq branches puis le consolidé. Une branche de plus, une
    de moins, et la lecture prendrait une branche pour l'ensemble sans qu'aucune
    somme ne bronche : les postes d'une branche font le total de cette branche.
    """
    lu = normaliser(_cellule(rang, colonne))
    if lu != attendu:
        raise BudgetSocialIncoherent(
            f"onglet {onglet} : la colonne {colonne} est intitulée « {lu} » et non"
            f" « {attendu} ». Les colonnes du tableau ont bougé — ce qui serait lu"
            " ne serait plus la grandeur attendue."
        )


def lire_equilibre(feuille) -> tuple[dict[tuple[str, str], dict[str, float]], dict, int]:
    """-> ({(section, intitulé): {branche ou 'consolide': montant en euros}},
           totaux de section et résultat net, nombre de lignes chiffrées lues).

    Le compte de lignes lues ignore la nomenclature : il compte tout intitulé
    suivi d'un nombre dans la colonne consolidée. C'est lui que la couverture
    compare à ce qui a été écrit, et c'est ce qui rend la comparaison honnête —
    un contrôle qui ne compterait que les lignes reconnues serait toujours vert.
    """
    section = ""
    lignes: dict[tuple[str, str], dict[str, float]] = {}
    reperes: dict[str, dict[str, float]] = {}
    lues = 0
    entete_vue = False
    for rang in feuille.iter_rows(values_only=True):
        intitule = normaliser(_cellule(rang, COLONNE_LIBELLE_EQUILIBRE))
        if not entete_vue and normaliser(_cellule(rang, COLONNE_CONSOLIDE)) == ENTETE_CONSOLIDE:
            controler_entete(rang, COLONNE_CONSOLIDE, ENTETE_CONSOLIDE, ONGLET_EQUILIBRE)
            for branche, colonne in BRANCHES_EQUILIBRE.items():
                controler_entete(rang, colonne, branche, ONGLET_EQUILIBRE)
            entete_vue = True
            continue
        consolide = _nombre(_cellule(rang, COLONNE_CONSOLIDE))
        if not intitule or consolide is None:
            continue
        lues += 1
        valeurs = {"consolide": consolide * MILLIONS}
        for branche, colonne in BRANCHES_EQUILIBRE.items():
            montant = _nombre(_cellule(rang, colonne))
            if montant is not None:
                valeurs[branche] = montant * MILLIONS
        if intitule in (SECTION_CHARGES, SECTION_PRODUITS):
            section = intitule
            reperes[intitule] = valeurs
            continue
        if intitule == LIGNE_RESULTAT:
            reperes[LIGNE_RESULTAT] = valeurs
            continue
        lignes[(section, intitule)] = valeurs
    if not entete_vue:
        raise BudgetSocialIncoherent(
            f"onglet {ONGLET_EQUILIBRE} : aucune colonne « {ENTETE_CONSOLIDE} »."
            " Le tableau des charges et produits nets a changé de forme."
        )
    manquants = [nom for nom in (SECTION_CHARGES, SECTION_PRODUITS, LIGNE_RESULTAT)
                 if nom not in reperes]
    if manquants:
        raise BudgetSocialIncoherent(
            f"onglet {ONGLET_EQUILIBRE} : lignes de total absentes ({', '.join(manquants)})."
        )
    return lignes, reperes, lues


def lire_soldes(feuille) -> dict[str, float]:
    """Le solde après mesures de chaque branche, tableau 4 de la même annexe."""
    entete_vue = False
    for rang in feuille.iter_rows(values_only=True):
        if not entete_vue and normaliser(
            _cellule(rang, COLONNE_SOLDES_CONSOLIDE)
        ) == ENTETE_SOLDES_CONSOLIDE:
            for branche, colonne in BRANCHES_SOLDES.items():
                controler_entete(rang, colonne, branche, ONGLET_SOLDES)
            entete_vue = True
            continue
        if normaliser(_cellule(rang, COLONNE_LIBELLE_SOLDES)) != LIGNE_SOLDE_APRES_MESURES:
            continue
        if not entete_vue:
            raise BudgetSocialIncoherent(
                f"onglet {ONGLET_SOLDES} : la ligne « {LIGNE_SOLDE_APRES_MESURES} »"
                " précède l'en-tête des branches."
            )
        soldes = {
            branche: (_nombre(_cellule(rang, colonne)) or 0.0) * MILLIONS
            for branche, colonne in BRANCHES_SOLDES.items()
        }
        soldes["consolide"] = (
            _nombre(_cellule(rang, COLONNE_SOLDES_CONSOLIDE)) or 0.0
        ) * MILLIONS
        return soldes
    raise BudgetSocialIncoherent(
        f"onglet {ONGLET_SOLDES} : pas de ligne « {LIGNE_SOLDE_APRES_MESURES} »."
        " Le second calcul du solde a disparu, et avec lui le seul contrôle qui"
        " ne partage aucun chiffre avec le premier."
    )


def lire_ondam(feuille) -> tuple[dict[str, dict[str, float]], int]:
    """-> ({code: {'base', 'taux', 'montant'}}, lignes chiffrées lues)."""
    mesures: dict[str, dict[str, float]] = {}
    lues = 0
    for rang in feuille.iter_rows(values_only=True):
        intitule = normaliser(_cellule(rang, COLONNE_LIBELLE_ONDAM))
        montant = _nombre(_cellule(rang, COLONNE_ONDAM_MONTANT))
        if not intitule or montant is None:
            continue
        lues += 1
        connu = SOUS_OBJECTIFS.get(intitule)
        if not connu:
            continue
        mesures[connu[0]] = {
            "base": (_nombre(_cellule(rang, COLONNE_ONDAM_BASE)) or 0.0) * MILLIARDS,
            "taux": _nombre(_cellule(rang, COLONNE_ONDAM_TAUX)) or 0.0,
            "montant": montant * MILLIARDS,
        }
    return mesures, lues


def lire_taux(feuille) -> list[float]:
    """Les taux d'évolution de la synthèse, dans l'ordre où elle les publie."""
    return [
        _nombre(_cellule(rang, COLONNE_TAUX_VALEUR))
        for rang in feuille.iter_rows(values_only=True)
        if normaliser(_cellule(rang, COLONNE_TAUX_LIBELLE))
        and _nombre(_cellule(rang, COLONNE_TAUX_VALEUR)) is not None
    ]


def montants(
    lignes: dict[tuple[str, str], dict[str, float]], perimetre: str = "consolide"
) -> dict[str, float]:
    """{code de poste: montant en euros} pour une colonne, selon la nomenclature."""
    return {
        POSTES[cle][0]: valeurs[perimetre]
        for cle, valeurs in lignes.items()
        if cle in POSTES and perimetre in valeurs
    }


def montants_par_branche(
    lignes: dict[tuple[str, str], dict[str, float]],
) -> dict[str, dict[str, float]]:
    """{clé de branche: {code de poste: montant en euros}}, les cinq branches."""
    return {
        CLE_BRANCHE[branche]: montants(lignes, branche) for branche in BRANCHES_EQUILIBRE
    }


def controler_sommes(postes: dict[str, float], reperes: dict, perimetre: str = "consolide") -> dict:
    """Les postes font leur total, à chaque étage, des deux côtés.

    Trois familles de sommes : chaque poste à enfants contre ses enfants, chaque
    section contre ses agrégats de tête, et le résultat net contre la différence
    des deux sections. Aucune ne se déduit d'une autre — un poste rattaché au
    mauvais parent casse la première sans toucher la deuxième, une ligne perdue
    au filtrage casse la deuxième sans toucher la première.

    `perimetre` nomme la colonne du tableau : le consolidé par défaut, ou une
    branche. Les trois familles de sommes valent pour chacune — c'est ce qui
    rend une branche réglable, et c'est vérifié branche par branche.
    """
    enfants: dict[str, list[str]] = defaultdict(list)
    tetes: dict[str, list[str]] = defaultdict(list)
    for (section, _), (code, parent, _) in POSTES.items():
        if parent:
            enfants[parent].append(code)
        else:
            tetes[section].append(code)
    ecarts = {}
    for parent, fils in sorted(enfants.items()):
        if parent not in postes:
            continue
        connus = [f for f in fils if f in postes]
        ecarts[parent] = round(sum(postes[f] for f in connus) - postes[parent], 2)
    for section in COTE:
        presents = [code for code in tetes[section] if code in postes]
        ecarts[section] = round(
            sum(postes[c] for c in presents) - reperes[section][perimetre], 2
        )
    ecarts[LIGNE_RESULTAT] = round(
        reperes[SECTION_PRODUITS][perimetre]
        - reperes[SECTION_CHARGES][perimetre]
        - reperes[LIGNE_RESULTAT][perimetre],
        2,
    )
    fautifs = {nom: ecart for nom, ecart in ecarts.items() if abs(ecart) > TOLERANCE_EUROS}
    if fautifs:
        nom, ecart = next(iter(sorted(fautifs.items())))
        raise BudgetSocialIncoherent(
            f"périmètre {perimetre} : « {nom} » s'écarte de {ecart / MILLIONS:+,.2f} M€"
            " de la somme de ses composantes. Un poste manque, double, ou n'est pas"
            " rattaché au bon parent : l'arbre publié ne serait pas le tableau lu."
        )
    return {
        "postes_verifies": len(ecarts),
        "ecart_maximal_euros": round(max(abs(e) for e in ecarts.values()), 2),
    }


def controler_les_branches(
    par_branche: dict[str, dict[str, float]], reperes: dict
) -> dict:
    """Ce qui autorise à régler une branche, et ce qui interdit de les additionner.

    **Chaque branche est un budget complet.** Ses postes font ses totaux, et son
    résultat net est la différence de ses deux totaux : c'est `controler_sommes`
    appliqué à sa colonne. Sans cela, une branche affichée serait un extrait, pas
    un budget, et la régler ne voudrait rien dire.

    **Les cinq résultats nets font le résultat consolidé, exactement.** C'est la
    prise décisive, et elle est contre-intuitive : les charges des cinq branches
    dépassent les charges consolidées de 19,0 Md€, et leurs produits les
    dépassent **du même montant**, parce que la consolidation retire un transfert
    entre branches des deux côtés à la fois. Les soldes, eux, ne bougent pas.
    C'est ce qui rend l'exercice honnête : **régler une branche déplace le solde
    de la Sécurité sociale d'exactement autant**, alors même que les totaux, eux,
    ne s'additionnent pas.

    **Et l'écart est non nul.** Si les cinq branches faisaient soudain le
    consolidé sur les charges, ce ne serait pas une bonne nouvelle : ce serait la
    preuve qu'on a lu cinq fois la même colonne. Le contrôle exige donc que
    l'écart existe, en plus d'être symétrique.
    """
    par_branche_verifie = {
        cle: controler_sommes(postes, reperes, branche)
        for branche, cle in CLE_BRANCHE.items()
        if (postes := par_branche.get(cle))
    }
    if len(par_branche_verifie) != len(CLE_BRANCHE):
        absentes = sorted(set(CLE_BRANCHE.values()) - set(par_branche_verifie))
        raise BudgetSocialIncoherent(
            f"branches sans aucun poste lu : {', '.join(absentes)}. Le tableau 5"
            " a changé de colonnes."
        )

    somme_soldes = sum(reperes[LIGNE_RESULTAT][b] for b in BRANCHES_EQUILIBRE)
    ecart_soldes = round(somme_soldes - reperes[LIGNE_RESULTAT]["consolide"], 2)
    if abs(ecart_soldes) > TOLERANCE_EUROS:
        raise BudgetSocialIncoherent(
            f"les cinq résultats nets font {ecart_soldes / MILLIONS:+,.2f} M€ de plus"
            " que le résultat consolidé. Régler une branche ne déplacerait alors pas"
            " le solde de la Sécurité sociale d'autant, et le simulateur mentirait."
        )

    ecarts_sections = {
        section: round(
            sum(reperes[section][b] for b in BRANCHES_EQUILIBRE) - reperes[section]["consolide"],
            2,
        )
        for section in (SECTION_CHARGES, SECTION_PRODUITS)
    }
    dissymetrie = round(ecarts_sections[SECTION_CHARGES] - ecarts_sections[SECTION_PRODUITS], 2)
    if abs(dissymetrie) > TOLERANCE_EUROS:
        raise BudgetSocialIncoherent(
            f"la consolidation retire {dissymetrie / MILLIONS:+,.2f} M€ de plus aux"
            " charges qu'aux produits. Un transfert entre branches se retire des deux"
            " côtés à la fois : la colonne consolidée ne décrit plus les mêmes lignes"
            " que les colonnes de branche."
        )
    if abs(ecarts_sections[SECTION_CHARGES]) <= TOLERANCE_EUROS:
        raise BudgetSocialIncoherent(
            "les cinq branches font exactement les charges consolidées. Il n'y aurait"
            " alors aucun transfert entre branches, ce que le tableau contredit :"
            " c'est la colonne consolidée qui a été lue cinq fois."
        )

    return {
        "branches_verifiees": sorted(par_branche_verifie),
        "postes_verifies_par_branche": {
            cle: v["postes_verifies"] for cle, v in sorted(par_branche_verifie.items())
        },
        "les_cinq_soldes_font_le_solde_consolide": ecart_soldes,
        "transferts_entre_branches_euros": round(ecarts_sections[SECTION_CHARGES], 2),
        "dissymetrie_euros": dissymetrie,
    }


def controler_deux_calculs_du_solde(reperes: dict, soldes: dict[str, float]) -> dict:
    """Le tableau 4 et le tableau 5 doivent donner le même solde, branche à branche.

    Le tableau 5 le pose comme produits moins charges ; le tableau 4 le construit
    comme un solde tendanciel corrigé de quarante-huit mesures. Les deux chemins
    ne partagent aucun calcul : c'est la seule prise de ce module qui ne puisse
    pas être satisfaite par une erreur de lecture systématique.
    """
    ecarts = {}
    for branche in (*BRANCHES_SOLDES, "consolide"):
        attendu = reperes[LIGNE_RESULTAT].get(branche)
        if attendu is None:
            continue
        ecarts[branche] = round(soldes[branche] - attendu, 2)
    if not ecarts:
        raise BudgetSocialIncoherent(
            "le tableau 5 ne donne aucun résultat net par branche : le contrôle"
            " croisé avec le tableau 4 n'a rien à comparer."
        )
    fautifs = {b: e for b, e in ecarts.items() if abs(e) > TOLERANCE_EUROS}
    if fautifs:
        branche, ecart = next(iter(sorted(fautifs.items())))
        raise BudgetSocialIncoherent(
            f"branche {branche} : le tableau 4 donne un solde après mesures qui"
            f" s'écarte de {ecart / MILLIONS:+,.2f} M€ du résultat net du tableau 5."
            " Les deux tableaux de la même annexe ne décrivent plus le même exercice."
        )
    return {
        "branches_verifiees": sorted(ecarts),
        "ecart_maximal_euros": round(max(abs(e) for e in ecarts.values()), 2),
    }


def controler_ondam(mesures: dict[str, dict[str, float]], taux_synthese: list[float]) -> dict:
    """Les six sous-objectifs font l'ONDAM, et la synthèse redonne les mêmes taux.

    Deux prises, sur deux onglets. La première somme les sous-objectifs votés —
    le regroupement médico-social ne compte pas, ses deux composantes le
    remplacent. La seconde reconstruit chaque montant depuis sa base et le taux
    publié *ailleurs dans le fichier*, où les intitulés ne sont pas les mêmes :
    le rapprochement se fait par position, et la position est déclarée.
    """
    manquants = [code for code in RANG_DES_TAUX if code not in mesures]
    if manquants:
        raise BudgetSocialIncoherent(
            f"ONDAM : sous-objectifs absents du tableau 8 ({', '.join(manquants)})."
            " Le périmètre voté a changé de découpage."
        )
    total = mesures["O-TOTAL"]["montant"]
    somme = sum(mesures[code]["montant"] for code in FEUILLES_OBJECTIF)
    if abs(somme - total) > TOLERANCE_ONDAM:
        raise BudgetSocialIncoherent(
            f"ONDAM : les six sous-objectifs font {somme / MILLIARDS:,.1f} Md€ contre"
            f" {total / MILLIARDS:,.1f} Md€ pour l'objectif total."
        )
    regroupements: dict[str, list[str]] = defaultdict(list)
    for code, parent, _ in SOUS_OBJECTIFS.values():
        if parent:
            regroupements[parent].append(code)
    for parent, fils in sorted(regroupements.items()):
        ecart = sum(mesures[f]["montant"] for f in fils) - mesures[parent]["montant"]
        if abs(ecart) > TOLERANCE_ONDAM:
            raise BudgetSocialIncoherent(
                f"{parent} : ses composantes s'en écartent de"
                f" {ecart / MILLIARDS:+,.2f} Md€. Le regroupement publié ne couvre"
                " plus les sous-objectifs qu'il annonce."
            )
    if len(taux_synthese) != len(RANG_DES_TAUX):
        raise BudgetSocialIncoherent(
            f"synthèse de l'annexe 5 : {len(taux_synthese)} taux publiés pour"
            f" {len(RANG_DES_TAUX)} sous-objectifs attendus. Le rapprochement se fait"
            " par position : une ligne de plus ou de moins le rendrait faux."
        )
    ecarts = {}
    for code, taux in zip(RANG_DES_TAUX, taux_synthese, strict=True):
        mesure = mesures[code]
        if abs(taux - mesure["taux"]) > 1e-9:
            raise BudgetSocialIncoherent(
                f"{code} : la synthèse publie un taux de {taux:.3f} là où le"
                f" tableau 8 en publie {mesure['taux']:.3f}. Les deux onglets ne"
                " décrivent plus le même sous-objectif au même rang."
            )
        ecarts[code] = round(mesure["base"] * (1 + taux) - mesure["montant"], 2)
    fautifs = {code: e for code, e in ecarts.items() if abs(e) > TOLERANCE_ONDAM}
    if fautifs:
        code, ecart = next(iter(sorted(fautifs.items())))
        raise BudgetSocialIncoherent(
            f"{code} : la base 2025 augmentée de son taux donne"
            f" {ecart / MILLIARDS:+,.2f} Md€ d'écart au montant 2026 publié."
        )
    return {
        "somme_des_sous_objectifs_euros": round(somme, 2),
        "ondam_total_euros": round(total, 2),
        "ecart_maximal_euros": round(max(abs(e) for e in ecarts.values()), 2),
    }


def controler_ordre_de_grandeur(conn, charges: float) -> dict:
    """Les charges des régimes de base, confrontées à une autre source déjà chargée.

    Tous les autres contrôles de ce module lisent le même fichier : ils resteraient
    verts si la DSS republiait ses cellules en milliards, et la Sécurité sociale
    entière s'afficherait divisée par mille. Celui-ci compare deux nombres qui
    viennent de deux producteurs — la DSS et la DREES — et il est le seul qu'un
    changement d'unité ne puisse pas satisfaire.

    La DREES absente de l'entrepôt ne vaut pas échec : elle vaut un contrôle non
    tenu, tracé comme tel plutôt que compté comme réussi.
    """
    rangs = conn.execute(
        """
        select period, value from core.observations
        where indicator_id = ? and geo_code = 'FR' and value is not null
        order by period desc limit 1
        """,
        (INDICATEUR_DREES,),
    ).fetchall()
    if not rangs:
        return {"tenu": False, "raison": f"{INDICATEUR_DREES} absent de l'entrepôt"}
    exercice, prestations = rangs[0]
    rapport = charges / float(prestations)
    bas, haut = RAPPORT_DREES
    if not bas <= rapport <= haut:
        raise BudgetSocialIncoherent(
            f"charges nettes des régimes de base {charges / MILLIONS:,.0f} M€ contre"
            f" {float(prestations) / MILLIONS:,.0f} M€ de prestations de protection"
            f" sociale (DREES {exercice}), soit un rapport de {rapport:.3f} hors de"
            f" [{bas}, {haut}]. Une unité ou une colonne a changé : les deux"
            " grandeurs ne peuvent pas s'éloigner à ce point."
        )
    return {
        "tenu": True,
        "rapport": round(rapport, 3),
        "prestations_drees_euros": float(prestations),
        "exercice_drees": str(exercice),
    }


def lignes_a_ecrire(
    exercice: str, postes: dict[str, float], mesures: dict[str, dict[str, float]]
) -> list[tuple]:
    """-> les lignes de `fin.social_budget_detail` pour un exercice.

    L'arbre y est porté par `parent_code`, comme celui du budget de l'État : un
    poste sans enfant est une feuille, quel que soit son étage.
    """
    annee = int(exercice)
    lignes = [
        (annee, LOI, COTE[section], niveau, code, parent, LIBELLES[code],
         round(postes[code], 2), 1, MESURE_EQUILIBRE)
        for (section, _), (code, parent, niveau) in POSTES.items()
        if code in postes
    ]
    lignes += [
        (annee, LOI, "objectif", niveau, code, parent, LIBELLES_OBJECTIF[code],
         round(mesures[code]["montant"], 2), 1, MESURE_OBJECTIF)
        for code, parent, niveau in SOUS_OBJECTIFS.values()
        if code in mesures
    ]
    return lignes


def lignes_branches_a_ecrire(
    exercice: str, par_branche: dict[str, dict[str, float]]
) -> list[tuple]:
    """-> les lignes de `fin.social_budget_branch`, cinq fois la nomenclature.

    Les codes sont ceux du consolidé, sans préfixe : c'est la colonne `branch`
    qui distingue, et la clé d'unicité de la table la porte. Un code préfixé
    aurait mis le périmètre dans l'identifiant d'une ligne, où il n'a rien à
    faire — et aurait fait diverger l'arbre d'une branche de celui du consolidé
    alors que c'est le même tableau.
    """
    annee = int(exercice)
    return [
        (annee, LOI, branche, COTE[section], niveau, code, parent, LIBELLES[code],
         round(postes[code], 2), 1, MESURE_EQUILIBRE)
        for branche, postes in sorted(par_branche.items())
        for (section, _), (code, parent, niveau) in POSTES.items()
        if code in postes
    ]


def ecrire_branches(conn, run_id: str, lignes: list[tuple]) -> int:
    """Remplace les cinq branches entières : un seul connecteur écrit ces lignes."""
    conn.execute("delete from fin.social_budget_branch where dataset_id = ?", (DATASET,))
    for ligne in lignes:
        conn.execute(
            """
            insert into fin.social_budget_branch
                (fiscal_year, law, branch, side, node_level, code, parent_code, label,
                 amount, sign, measure, dataset_id, run_id)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*ligne, DATASET, run_id),
        )
    return len(lignes)


def ecrire(conn, run_id: str, lignes: list[tuple]) -> int:
    """Remplace la nomenclature entière : un seul connecteur écrit ces lignes."""
    conn.execute("delete from fin.social_budget_detail where dataset_id = ?", (DATASET,))
    for ligne in lignes:
        conn.execute(
            """
            insert into fin.social_budget_detail
                (fiscal_year, law, side, node_level, code, parent_code, label,
                 amount, sign, measure, dataset_id, run_id)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*ligne, DATASET, run_id),
        )
    return len(lignes)


def collecter(conn, store, run_id: str, exercice: str) -> dict:
    """Les deux annexes de l'exercice, archivées avant lecture."""
    # securite-sociale.fr ne négocie qu'une suite TLS sans confidentialité
    # persistante, qu'OpenSSL refuse à son niveau de sécurité par défaut : sans
    # ce contexte, le téléchargement échoue en CI sur une poignée de main. La
    # chaîne de certification, elle, reste vérifiée — voir le module `http`.
    contexte = contexte_sans_confidentialite_persistante()
    classeurs = {}
    for annexe in ("equilibre", "ondam"):
        adresse = url(exercice, annexe)
        contenu = telecharger(adresse, timeout=300, verify=contexte)
        entrepot.record_asset(
            conn, store, run_id, DATASET, SOURCE, ANNEXES[exercice][annexe], contenu,
            adresse,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        classeurs[annexe] = classeur(contenu)
    return classeurs


def preparer(conn, classeurs: dict) -> tuple[dict, dict, dict, dict]:
    """-> (postes consolidés, mesures ONDAM, contrôles, comptes lus/écrits)."""
    lignes, reperes, lues_equilibre = lire_equilibre(classeurs["equilibre"][ONGLET_EQUILIBRE])
    soldes = lire_soldes(classeurs["equilibre"][ONGLET_SOLDES])
    mesures, lues_ondam = lire_ondam(classeurs["ondam"][ONGLET_ONDAM])
    taux = lire_taux(classeurs["ondam"][ONGLET_TAUX])
    postes = montants(lignes)
    par_branche = montants_par_branche(lignes)
    verifies = {
        "les_postes_font_leur_total": controler_sommes(postes, reperes),
        "chaque_branche_est_un_budget_complet": controler_les_branches(par_branche, reperes),
        "deux_tableaux_donnent_le_meme_solde": controler_deux_calculs_du_solde(reperes, soldes),
        "les_sous_objectifs_font_l_ondam": controler_ondam(mesures, taux),
        "ordre_de_grandeur_contre_la_drees": controler_ordre_de_grandeur(
            conn, reperes[SECTION_CHARGES]["consolide"]
        ),
        "charges_nettes_euros": round(reperes[SECTION_CHARGES]["consolide"], 2),
        "produits_nets_euros": round(reperes[SECTION_PRODUITS]["consolide"], 2),
        "resultat_net_euros": round(reperes[LIGNE_RESULTAT]["consolide"], 2),
    }
    # Les deux totaux de section et le résultat net sont lus mais pas écrits :
    # le site les recalcule depuis les feuilles, comme il le fait déjà pour le
    # budget de l'État. Ils sortent donc du dénominateur de la couverture.
    comptes = {
        "equilibre": (lues_equilibre - 3, len(postes)),
        "ondam": (lues_ondam, len(mesures)),
    }
    return postes, par_branche, mesures, verifies, comptes


def run(store_spec: str) -> int:
    conn = entrepot.connect()
    store = make_store(store_spec)
    run_id = entrepot.start_run(conn, DATASET, "manual")
    try:
        garde_fou_volume(conn)
        ecrites_total = 0
        verifies_par_exercice = {}
        for exercice in sorted(ANNEXES):
            classeurs = collecter(conn, store, run_id, exercice)
            postes, par_branche, mesures, verifies, comptes = preparer(conn, classeurs)
            # Ce qui est écrit contre ce qui est lu : une ligne que la DSS
            # ajouterait au tableau ne serait pas dans la nomenclature, donc pas
            # écrite, et la part reconnue tomberait sous le seuil. Les contrôles
            # de somme, eux, ne verraient rien — un détail supplémentaire ne
            # change aucun total.
            parts = couverture.controler(
                {nom: lu for nom, (lu, _) in comptes.items()},
                {nom: ecrit for nom, (_, ecrit) in comptes.items()},
            )
            verifies["couverture_lu_ecrit"] = {nom: round(part, 3) for nom, part in parts.items()}
            ecrites = ecrire(conn, run_id, lignes_a_ecrire(exercice, postes, mesures))
            branches = ecrire_branches(
                conn, run_id, lignes_branches_a_ecrire(exercice, par_branche)
            )
            ecrites_total += ecrites + branches
            verifies_par_exercice[exercice] = verifies
            entrepot.etape(
                f"{exercice} : {ecrites} postes écrits + {branches} par branche,"
                f" {verifies['charges_nettes_euros'] / 1e9:.1f} Md€ de charges nettes"
            )
        for nom in (
            "les_postes_font_leur_total",
            "chaque_branche_est_un_budget_complet",
            "deux_tableaux_donnent_le_meme_solde",
            "les_sous_objectifs_font_l_ondam",
            "ordre_de_grandeur_contre_la_drees",
            "couverture_lu_ecrit",
        ):
            conn.execute(
                """
                insert into meta.data_quality_checks
                    (run_id, dataset_id, check_name, severity, passed, observed)
                values (?, ?, ?, 'blocker', true, ?)
                """,
                (run_id, DATASET, nom,
                 json.dumps({e: v[nom] for e, v in verifies_par_exercice.items()})),
            )
        conn.commit()
        entrepot.finish_run(conn, run_id, "success", rows_written=ecrites_total)
        dernier = max(verifies_par_exercice)
        resume = verifies_par_exercice[dernier]
        print(
            f"PLFSS {dernier} : {ecrites_total} postes,"
            f" {resume['charges_nettes_euros'] / 1e9:.1f} Md€ de charges nettes,"
            f" {resume['produits_nets_euros'] / 1e9:.1f} Md€ de produits nets,"
            f" résultat {resume['resultat_net_euros'] / 1e9:+.1f} Md€ ;"
            f" ONDAM {resume['les_sous_objectifs_font_l_ondam']['ondam_total_euros'] / 1e9:.1f}"
            " Md€ en six sous-objectifs"
        )
        return 0
    except Exception as error:  # noqa: BLE001 — tout échec finit tracé dans le lineage
        entrepot.finish_run(conn, run_id, "failed", error=str(error))
        raise
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default=".snapshots")
    return run(parser.parse_args().store)


if __name__ == "__main__":
    raise SystemExit(main())
