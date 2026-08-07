"""Fonds européens conventionnés par région : FEDER, FSE+ et FTJ 2021-2027 (ANCT).

7,88 milliards d'euros d'aide européenne engagée sur 16 625 opérations, au
16 mars 2026. C'est la liste que le règlement européen oblige chaque autorité de
gestion à publier : un projet, un bénéficiaire, un montant.

**Conventionné n'est pas payé, et c'est la réserve qui commande tout le
reste.** Ce montant est celui qu'une convention engage ; il sera versé au fil
des justificatifs, sur des années, et une partie ne le sera jamais — projets
abandonnés, dépenses rejetées au contrôle, sous-réalisation. Les paiements
n'existent nulle part dans ce fichier : les publier demanderait une autre
source, et le seul état de paiement diffusé par l'ANCT est un PDF graphique
qu'aucune machine ne lit. Présenter ces 7,88 Md€ comme des versements serait le
contresens que ce module existe pour empêcher.

**La colonne géographique évidente est inutilisable.** `Région de l'opération`
est vide sur 6 135 lignes (36,9 %) et porte 71 valeurs distinctes pour
18 régions : « 24/Centre-Val de Loire », « | 28/Normandie », « Hauts-de-France »,
« La Réunion » et « 04/La Réunion » cohabitent. `Département de l'opération` est
vide à 64,7 %. La seule variable régionale fiable est `Libellé Programme` :
21 modalités, aucune manquante, une par autorité de gestion. La régionalisation
passe donc par le programme, et la correspondance vers les codes INSEE est
écrite ici plutôt que devinée à l'exécution.

**Ce qui n'est pas régionalisé est dit, pas réparti.** Les deux programmes
nationaux — FSE+ Emploi-Inclusion-Jeunesse-Compétences et FTJ Emploi-Compétences
— pèsent 2,32 Md€ et n'appartiennent à aucune région ; le programme de
Saint-Martin (14,1 M€) relève d'une collectivité d'outre-mer, absente du
référentiel régional. Les 398 opérations que la source marque « Volet national »
(651,1 M€) ne sont pas localisables non plus. Rien de tout cela n'est ventilé au
prorata de quoi que ce soit : 5,54 Md€ portent une région, le reste est chiffré
dans le contrôle qualité du run.

**Les deux contrôles qui bloquent.** `Total des dépenses éligibles` ×
`Taux de cofinancement` doit redonner `Montant UE` sur les 16 591 lignes qui
portent les trois montants — aucun écart ne dépasse 1 %, et un décalage de
colonne le ferait exploser. Et le décompte doit valoir exactement les 16 625
opérations de la parution chargée : c'est le seul garde-fou contre une lecture
tronquée, un classeur qui ne rendrait qu'une partie de sa feuille ne se voyant
sur aucun total interne.

**L'URL du fichier change à chaque parution** : elle porte la date de
publication. Elle est retrouvée sur la page de référence, comme `elections.py`
retrouve la sienne par le titre dans le catalogue de data.gouv.

Usage : python -m plateforme.normalize.fonds_europeens [--store r2:plateforme-raw]
"""

import argparse
import io
import json
import math
import re
from collections import defaultdict

from plateforme import couverture, entrepot, revisions
from plateforme.http import fetch, telecharger
from plateforme.limites import garde_fou_volume
from plateforme.normalize.geo import MILLESIME, make_store
from plateforme.normalize.ofgl import filtrer_territoires_connus

DATASET = "anct-fonds-europeens-2021-2027"
SOURCE = "europe-en-france"
PAGE = (
    "https://www.europe-en-france.gouv.fr/fr/ressources/"
    "liste-operations-feder-fse-ftj-2021-2027"
)
# Le nom du fichier commence par sa date de publication : 20260316_liste_…xlsx.
# C'est ce préfixe qui départage deux parutions laissées côte à côte sur la page.
LIEN = re.compile(
    r'href="(https://[^"]+/(\d{8})_liste_operations_conventionnees[^"]*\.xlsx)"'
)
# La feuille des données, et la date à laquelle la liste est arrêtée.
FEUILLE = re.compile(r"LISTE OPERATION AU (\d{2}) (\d{2}) (\d{4})")

# Les huit colonnes lues sur les vingt-trois du fichier. Les autres décrivent le
# projet (intitulé, objectifs, bénéficiaire, code postal, dates) ; les nommer
# vaut refus explicite de les charger. `Région de l'opération` n'est pas lue pour
# localiser — elle ne le peut pas — mais pour reconnaître le « Volet national ».
COLONNES_LUES = {
    "operation": "Numéro Opération",
    "cci": "NUMCCI",
    "programme": "Libellé Programme",
    "region_source": "Région de l'opération",
    "fonds": "Fonds",
    "depenses": "Total des dépenses éligibles",
    "taux": "Taux de cofinancement",
    "ue": "Montant UE",
}
MONTANTS = ("depenses", "taux", "ue")

# Programme -> code région INSEE, ou None pour ce qui n'est pas une région. La
# correspondance est écrite plutôt que déduite du libellé : « Grand Est et massif
# des Vosges » ou « Centre-Val de Loire et interrégional Loire » ne se laissent
# pas deviner, et un programme inconnu doit faire échouer le run plutôt que
# disparaître silencieusement de la carte. Les libellés sont comparés après
# normalisation des espaces : la source écrit une espace insécable dans « Pays de
# la Loire » et double l'espace dans « Mayotte » et « Occitanie ».
PROGRAMMES = {
    "Programme Bretagne FEDER-FSE+ 2021-2027": "53",
    "Programme Centre-Val de Loire et interrégional Loire FEDER-FSE+ 2021-2027": "24",
    "Programme Corse FEDER-FSE+ 2021-2027": "94",
    "Programme FEDER FSE+ Nouvelle-Aquitaine": "75",
    "Programme FEDER-FSE+ 2021-2027 de La Réunion": "04",
    "Programme FEDER-FSE+ Auvergne-Rhône-Alpes 2021-2027": "84",
    "Programme FEDER-FSE+ Bourgogne Franche-Comté et Massif du Jura 2021-2027": "27",
    "Programme FEDER-FSE+ GUYANE 2021-2027": "03",
    "Programme FEDER-FSE+ Guadeloupe 2021-2027": "01",
    "Programme Grand Est et massif des Vosges FEDER-FSE+-FTJ 2021-2027": "44",
    "Programme Hauts de France FEDER-FSE+-FTJ 2021-2027": "32",
    "Programme Martinique FEDER-FSE+ 2021-2027": "02",
    "Programme Mayotte FEDER-FSE+ 2021- 2027": "06",
    "Programme Normandie FEDER-FSE+-FTJ 2021-2027": "28",
    "Programme Occitanie FEDER-FSE+ 2021- 2027": "76",
    "Programme Pays de la Loire FEDER-FSE+-FTJ 2021-2027": "52",
    "Programme Provence-Alpes-Côte d’Azur et Massif des Alpes FEDER-FSE+-FTJ 2021-2027": "93",
    "Programme régional Île-de-France et bassin de la Seine FEDER-FSE+ 2021-2027": "11",
    # Saint-Martin est une collectivité d'outre-mer : elle a un programme
    # européen et pas de code région au Code officiel géographique.
    "Programme Saint Martin FEDER 2021-2027": None,
    "Programme national FSE+ Emploi - Inclusion - Jeunesse - Compétences": None,
    "Programme national FTJ Emploi - Compétences": None,
}

# La valeur par laquelle la source marque les opérations qui ne relèvent d'aucun
# territoire. Elles sont toutes, dans cette parution, portées par le programme
# national — donc déjà hors maille régionale ; le filtre reste, pour le jour où
# un programme régional en portera une.
VOLET_NATIONAL = "Volet national"

INDICATEUR = "anct_fonds_europeens_ue_conventionne"
THEME = "europe"

# Une programmation, pas un exercice : les montants s'accumulent depuis 2021 et
# la période n'a pas d'autre borne que celle du règlement européen. La date à
# laquelle la liste est arrêtée n'entre pas dans la période — elle bougerait à
# chaque parution et fabriquerait une fausse série — mais dans le contrôle
# qualité du run, où elle se lit.
PERIODE = "2021-2027"

# Le décompte de la parution chargée. Ce n'est pas une invariante : la prochaine
# liste en comptera davantage. C'est un sceau — une lecture tronquée ne se voit
# sur aucun total interne, puisque la somme de ce qu'on a lu est toujours égale à
# elle-même. Le relever demande d'avoir regardé la nouvelle parution.
OPERATIONS = 16_625

# Écart relatif toléré entre le montant UE publié et son recalcul. Un pour cent
# laisse passer l'arrondi au centime et rien d'autre : sur la parution du
# 16 mars 2026, l'écart maximum est de 0,0011 % — et il vient d'une opération de
# 217,74 €, où un demi-centime pèse déjà cela. Une colonne qui aurait glissé
# ferait un écart de plusieurs dizaines de pour cent.
TOLERANCE = 0.01

FICHE = {
    "libelle": "Fonds européens conventionnés (FEDER, FSE+, FTJ) 2021-2027",
    "public": "Ce que l'Union européenne s'est engagée à verser dans la région"
    " au titre des projets déjà conventionnés de la programmation 2021-2027."
    " C'est un montant programmé, pas un montant payé : les versements"
    " effectifs ne figurent pas dans la source.",
    "technique": "Somme des montants d'aide européenne (FEDER, FSE+, FTJ) des"
    " opérations conventionnées de la programmation 2021-2027, source ANCT,"
    " autorité de coordination des fonds européens en France, en euros courants"
    " cumulés depuis 2021 et arrêtés à la date de publication de la liste."
    " **Ce sont des montants conventionnés, c'est-à-dire programmés, et non des"
    " montants payés** : l'aide est versée ensuite au fil des justificatifs, sur"
    " plusieurs années, et une part ne l'est jamais — projets abandonnés,"
    " dépenses rejetées au contrôle, sous-réalisation. Les paiements ne figurent"
    " pas dans cette source. La région est celle du programme opérationnel"
    " (colonne « Libellé Programme »), seule variable régionale complète du"
    " fichier : la colonne « Région de l'opération » est vide sur 36,9 % des"
    " lignes et porte 71 orthographes pour 18 régions, celle du département sur"
    " 64,7 %. Ne sont pas régionalisés, et ne sont donc répartis sur aucune"
    " région : les deux programmes nationaux (FSE+ Emploi-Inclusion-Jeunesse-"
    "Compétences et FTJ Emploi-Compétences, 2,32 Md€), le programme de"
    " Saint-Martin, collectivité d'outre-mer sans code région (14,1 M€), et les"
    " opérations que la source marque « Volet national » (398 opérations,"
    " 651,1 M€). Sur 7,88 Md€ conventionnés, 5,54 Md€ portent une région. Le"
    " montant régional n'est pas une dépense annuelle et ne se compare pas au"
    " budget d'un exercice ; il ne dit pas non plus ce que la région perçoit en"
    " tant que collectivité, l'autorité de gestion reversant l'aide aux"
    " bénéficiaires des projets. Le numéro d'opération n'identifie pas une ligne"
    " à lui seul — le même numéro sert dans deux programmes — et 34 opérations"
    " portent un montant UE nul ou absent.",
    "formule": "Somme des montants UE conventionnés des opérations du programme régional",
}


class ConventionnementRompu(RuntimeError):
    """Le fichier ne dit plus ce que la parution précédente disait."""


def _normaliser(texte: str) -> str:
    """Espaces insécables et doubles espaces ramenés à une espace simple."""
    return " ".join(str(texte).replace("\xa0", " ").split())


def trouver_url(page: str) -> str:
    """L'URL du classeur, retrouvée dans la page de référence.

    L'URL porte la date de parution et change à chaque mise à jour ; la page,
    elle, est stable. La chercher plutôt que la coder en dur évite de recharger
    éternellement une liste périmée sans que rien ne le dise.
    """
    liens = LIEN.findall(page)
    if not liens:
        raise ValueError(
            f"aucun fichier « …_liste_operations_conventionnees….xlsx » sur {PAGE} :"
            " l'ANCT a renommé son fichier ou déplacé la page, il faut regarder"
            " avant de recharger"
        )
    adresse, _ = max(liens, key=lambda lien: lien[1])
    return adresse


def _nombre(valeur) -> float | None:
    """Un montant du classeur -> float, ou None s'il n'est pas renseigné."""
    if valeur is None:
        return None
    if isinstance(valeur, (int, float)):
        return float(valeur)
    texte = _normaliser(valeur).replace(" ", "").replace(",", ".")
    return float(texte) if texte else None


def lire(contenu: bytes) -> tuple[list[dict], str]:
    """Classeur de l'ANCT -> (opérations, date d'arrêté).

    Les repères sont contrôlés plutôt que supposés : la feuille est celle dont
    le nom dit « LISTE OPERATION AU jj mm aaaa » — le classeur en porte une
    seconde, « INFO », qui ne contient que la notice — et les huit colonnes lues
    sont retrouvées par leur intitulé exact. Les montants sont arrondis au
    centime dès la lecture : le tableur rend 49999,99999999999 là où la
    convention dit 50 000, et un total de seize mille lignes de ce bruit ne
    tomberait jamais rond.
    """
    import openpyxl

    livre = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    feuilles = [(nom, FEUILLE.search(nom)) for nom in livre.sheetnames]
    feuilles = [(nom, trouve) for nom, trouve in feuilles if trouve]
    if len(feuilles) != 1:
        raise ValueError(
            "une seule feuille « LISTE OPERATION AU jj mm aaaa » est attendue,"
            f" {len(feuilles)} trouvée(s) parmi {livre.sheetnames} :"
            " le classeur a changé de forme"
        )
    nom, trouve = feuilles[0]
    jour, mois, annee = trouve.groups()
    arrete = f"{annee}-{mois}-{jour}"

    rangees = livre[nom].iter_rows(values_only=True)
    entete = {_normaliser(cellule or ""): rang for rang, cellule in enumerate(next(rangees))}
    manquantes = [nom for nom in COLONNES_LUES.values() if nom not in entete]
    if manquantes:
        raise ValueError(f"colonnes absentes du classeur : {manquantes}")

    operations = []
    for rangee in rangees:
        if not any(cellule is not None and str(cellule).strip() for cellule in rangee):
            continue
        operation = {
            cle: rangee[entete[colonne]] for cle, colonne in COLONNES_LUES.items()
        }
        for cle in MONTANTS:
            operation[cle] = _nombre(operation[cle])
        for cle in ("depenses", "ue"):
            if operation[cle] is not None:
                operation[cle] = round(operation[cle], 2)
        for cle in ("programme", "region_source", "operation", "cci", "fonds"):
            operation[cle] = _normaliser(operation[cle] or "")
        operations.append(operation)
    return operations, arrete


def montant_par_fonds(operations: list[dict]) -> dict[str, float]:
    """{fonds: montant UE conventionné}. Ni publié ni cartographié : tracé dans
    le contrôle qualité, où il dit d'où vient la masse."""
    par_fonds: dict[str, list[float]] = defaultdict(list)
    for operation in operations:
        par_fonds[operation["fonds"]].append(operation["ue"] or 0.0)
    return {
        fonds: round(math.fsum(montants), 2)
        for fonds, montants in sorted(par_fonds.items())
    }


def controler(operations: list[dict], attendu: int = OPERATIONS) -> dict:
    """Les deux contrôles qui bloquent, plus la complétude de la correspondance.

    Le premier — dépenses éligibles × taux de cofinancement = montant UE — est
    le seul qui puisse voir une colonne qui aurait glissé : il compare trois
    colonnes entre elles, là où un total ne se compare qu'à lui-même. Le second
    est le décompte : un classeur lu à moitié rend une somme parfaitement
    cohérente et parfaitement fausse.
    """
    if len(operations) != attendu:
        raise ConventionnementRompu(
            f"{len(operations)} opérations lues contre {attendu} attendues."
            " Lecture tronquée, ou nouvelle parution : dans le second cas, il faut"
            " avoir regardé la liste avant de relever le décompte attendu."
        )
    inconnus = sorted({operation["programme"] for operation in operations} - set(PROGRAMMES))
    if inconnus:
        raise ConventionnementRompu(
            f"programmes que la correspondance ne connaît pas : {', '.join(inconnus)}."
            " Sans code région, leurs opérations disparaîtraient de la carte sans"
            " manquer à aucun total."
        )
    verifiees, pire = 0, 0.0
    for operation in operations:
        depenses, taux, ue = operation["depenses"], operation["taux"], operation["ue"]
        if not ue or depenses is None or taux is None:
            continue
        verifiees += 1
        ecart = abs(depenses * taux - ue) / abs(ue)
        pire = max(pire, ecart)
        if ecart > TOLERANCE:
            raise ConventionnementRompu(
                f"opération {operation['cci']}/{operation['operation']} :"
                f" {depenses:,.2f} € × {taux:.4f} = {depenses * taux:,.2f} € contre"
                f" {ue:,.2f} € publiés, soit {100 * ecart:.2f} % d'écart. Les trois"
                " colonnes de montants ne décrivent plus la même opération."
            )
    if not verifiees:
        raise ConventionnementRompu(
            "aucune opération ne porte à la fois les dépenses, le taux et le"
            " montant UE : le contrôle ne vérifie plus rien"
        )
    return {
        "operations": len(operations),
        "lignes_verifiees": verifiees,
        "ecart_relatif_maximal": pire,
        "montant_ue_absent": sum(1 for operation in operations if not operation["ue"]),
        # `fsum` et non `sum` : seize mille additions flottantes dérivent de
        # quelques centimes sur un total de sept milliards, et un total publié au
        # centime doit être celui des centimes lus.
        "montant_ue_total": round(
            math.fsum(operation["ue"] or 0.0 for operation in operations), 2
        ),
        "depenses_eligibles_total": round(
            math.fsum(operation["depenses"] or 0.0 for operation in operations), 2
        ),
        "programmes": len({operation["programme"] for operation in operations}),
        # Trois fonds, trois logiques : le FEDER finance des investissements, le
        # FSE+ des actions pour l'emploi, le FTJ la sortie du charbon et du
        # carbone. Le total les mélange ; la ventilation reste tracée.
        "montant_ue_par_fonds": montant_par_fonds(operations),
        "volet_national": sum(
            1 for operation in operations if operation["region_source"] == VOLET_NATIONAL
        ),
        "cles_dupliquees": len(operations)
        - len({(operation["cci"], operation["operation"]) for operation in operations}),
    }


def par_region(operations: list[dict]) -> list[tuple]:
    """-> [(indicateur, niveau, code région, période, montant UE conventionné)].

    Le programme donne la région ; ce qui n'en a pas est écarté, jamais réparti.
    """
    totaux: dict[str, list[float]] = defaultdict(list)
    for operation in operations:
        code = PROGRAMMES.get(operation["programme"])
        if code is None or operation["region_source"] == VOLET_NATIONAL:
            continue
        totaux[code].append(operation["ue"] or 0.0)
    return [
        (INDICATEUR, "region", code, PERIODE, round(math.fsum(montants), 2))
        for code, montants in sorted(totaux.items())
    ]


def non_regionalise(operations: list[dict]) -> dict[str, dict]:
    """Ce qui reste hors de la maille régionale, chiffré motif par motif.

    Une couverture incomplète est une information à publier, pas à taire : ces
    2,34 Md€ n'ont pas de région, et le contrôle qualité du run les nomme.
    """
    resume: dict[str, list[float]] = defaultdict(list)
    for operation in operations:
        if PROGRAMMES.get(operation["programme"]) is None:
            motif = operation["programme"]
        elif operation["region_source"] == VOLET_NATIONAL:
            motif = f"volet national dans {operation['programme']}"
        else:
            continue
        resume[motif].append(operation["ue"] or 0.0)
    return {
        motif: {"operations": len(montants), "montant_ue": round(math.fsum(montants), 2)}
        for motif, montants in sorted(resume.items())
    }


def montant_par_maille(candidates) -> dict[str, float]:
    """Les euros écrits par maille : les deux côtés de la couverture.

    La couverture se mesure en euros et non en codes — une région manquante ne
    pèse pas le dix-huitième du total, elle pèse ce qu'elle a conventionné.
    """
    par_maille: dict[str, list[float]] = defaultdict(list)
    for _, niveau, _, _, montant in candidates:
        par_maille[niveau].append(montant)
    return {niveau: round(math.fsum(montants), 2) for niveau, montants in par_maille.items()}


def declarer(conn) -> None:
    with conn.cursor() as curseur:
        definition = curseur.execute(
            """
            insert into core.indicator_definitions
                (public_definition, technical_definition, formula, unit_notes,
                 confidence_level, badges)
            values (?, ?, ?, 'euros courants, montants conventionnés et non payés',
                    'observed', array['Officiel','Donnée brute'])
            returning definition_id
            """,
            (FICHE["public"], FICHE["technique"], FICHE["formule"]),
        ).fetchone()[0]
        # Pas de `accounting_frame` : une convention européenne n'est ni une
        # comptabilité budgétaire, ni générale, ni nationale — c'est un
        # engagement pluriannuel, et les trois cadres du schéma mentiraient.
        curseur.execute(
            """
            insert into core.indicators
                (indicator_id, dataset_id, definition_id, theme, label_fr, unit,
                 additive, price_basis, geo_levels, time_granularity, published)
            values (?, ?, ?, ?, ?, 'EUR', true, 'current', array['region'],
                    'pluriannuelle', true)
            on conflict (indicator_id) do update set unit = excluded.unit,
                definition_id = excluded.definition_id,
                label_fr = excluded.label_fr,
                theme = excluded.theme, published = true
            """,
            (INDICATEUR, DATASET, definition, THEME, FICHE["libelle"]),
        )
        curseur.execute(
            "delete from core.indicator_definitions d where not exists"
            " (select 1 from core.indicators i where i.definition_id = d.definition_id)"
        )
    conn.commit()


COLONNES = ("value",)


def run(store_spec: str) -> int:
    conn = entrepot.connect()
    store = make_store(store_spec)
    declarer(conn)
    run_id = entrepot.start_run(conn, DATASET, "manual")
    try:
        garde_fou_volume(conn)
        adresse = trouver_url(fetch(PAGE, timeout=120).text)
        entrepot.etape(f"fichier retrouvé sur la page de référence : {adresse}")
        contenu = telecharger(adresse, timeout=900)
        entrepot.record_asset(
            conn, store, run_id, DATASET, SOURCE,
            "liste-operations-conventionnees.xlsx", contenu, adresse,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        operations, arrete = lire(contenu)
        entrepot.etape(f"{len(operations)} opérations lues, liste arrêtée au {arrete}")
        verifies = controler(operations)
        entrepot.etape(
            f"{verifies['lignes_verifiees']} lignes où dépenses × taux redonne le"
            f" montant UE, écart maximal {100 * verifies['ecart_relatif_maximal']:.4f} %"
        )

        candidates = par_region(operations)
        gardees, ecartes = filtrer_territoires_connus(conn, candidates)
        ecrites, revisees = revisions.remplacer(
            conn, run_id, [INDICATEUR], COLONNES,
            [
                (identifiant, niveau, code, MILLESIME, periode, montant)
                for identifiant, niveau, code, periode, montant in gardees
            ],
        )
        lus = montant_par_maille(candidates)
        parts = couverture.controler(lus, montant_par_maille(gardees))
        conn.execute(
            """
            insert into meta.data_quality_checks
                (run_id, dataset_id, check_name, severity, passed, observed)
            values (?, ?, 'le_montant_ue_se_recalcule_et_le_decompte_est_complet',
                    'blocker', true, ?)
            """,
            (run_id, DATASET, json.dumps({
                **verifies,
                "arrete_au": arrete,
                "montant_ue_regionalise": lus.get("region", 0.0),
                "non_regionalise": non_regionalise(operations),
                "hors_referentiel": sorted(ecartes),
                "couverture": {n: round(p, 4) for n, p in sorted(parts.items())},
            }, ensure_ascii=False)),
        )
        conn.commit()
        entrepot.finish_run(
            conn, run_id, "success", rows_read=len(operations), rows_written=ecrites
        )
        regionalise = lus.get("region", 0.0)
        print(
            f"fonds européens : {ecrites} régions, {verifies['operations']} opérations"
            f" conventionnées au {arrete} pour"
            f" {verifies['montant_ue_total'] / 1e9:.2f} Md€ d'aide UE, dont"
            f" {regionalise / 1e9:.2f} Md€ régionalisés — montants conventionnés,"
            f" pas payés ; {revisees} révisées, {len(ecartes)} hors référentiel"
        )
        return 0
    except Exception as error:  # noqa: BLE001 — tout échec finit tracé dans le lineage
        entrepot.finish_run(conn, run_id, "failed", error=str(error))
        raise
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--store", default=".snapshots")
    return run(parser.parse_args().store)


if __name__ == "__main__":
    raise SystemExit(main())
